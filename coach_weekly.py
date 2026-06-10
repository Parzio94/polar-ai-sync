#!/usr/bin/env python3
"""
coach_weekly.py — Moteur d'analyse Coach IA élite
Polar App — Mattéo

Usage :
  python3 coach_weekly.py                    # analyse semaine courante
  python3 coach_weekly.py --week 2026-03-16  # analyse une semaine précise
  python3 coach_weekly.py --dry-run          # affiche le prompt sans appeler l'API

Sortie :
  /root/polar/analysis_current.json          # rapport coach (écrasé à chaque analyse)
  /root/polar/programme/semaine_XX.json      # programme semaine suivante mis à jour
  /root/polar/races.json                     # sub cibles + probabilités mis à jour
  /root/polar/hr_zones.json                  # zones mises à jour si confiance élevée
"""

import os, json, sys, math, re, argparse
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

# ─── CONFIGURATION ────────────────────────────────────────────────────────────

POLAR        = Path(os.getenv("KDRIVE_PATH", "/root/polar"))
DATA         = POLAR / "data"
WEEKS_DIR    = DATA / "weeks"
PROG_DIR     = POLAR / "programme"
ZONES_FILE   = POLAR / "hr_zones.json"
RACES_FILE   = POLAR / "races.json"
CAL_FILE     = POLAR / "calendar_events.json"
VOYAGES_FILE = POLAR / "voyages.json"
ANALYSIS_FILE = POLAR / "analysis_current.json"

ANTHROPIC_KEY = os.getenv("ANTHROPIC_API_KEY", "")

# Profil statique Mattéo
PROFIL = {
    "prenom":    "Mattéo",
    "age":       24,
    "poids_kg":  69,
    "taille_cm": 178,
}

# Seuils Option C pour mise à jour automatique des zones
ZONE_UPDATE_MIN_SEANCES  = 2    # ≥2 séances calibrantes
ZONE_UPDATE_MIN_BPM_DIFF = 3    # divergence >3 bpm
ZONE_UPDATE_MIN_PACE_DIFF = 8   # divergence >8 sec/km
ZONE_UPDATE_MAX_HR_DRIFT  = 8   # FC ne doit pas dériver >8 bpm sur séance longue Z2


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def parse_duration_s(s):
    """Convertit durée string en secondes."""
    if not s: return 0
    s = str(s)
    m = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:([\d.]+)S)?', s)
    if m:
        return int(m.group(1) or 0)*3600 + int(m.group(2) or 0)*60 + float(m.group(3) or 0)
    total = 0
    mh = re.search(r'(\d+)h', s);   total += int(mh.group(1))*3600 if mh else 0
    mm = re.search(r'(\d+)min', s); total += int(mm.group(1))*60   if mm else 0
    ms = re.search(r'(\d+)s', s);   total += int(ms.group(1))      if ms else 0
    if total: return total
    try: return int(float(s))
    except: return 0

def pace_str_to_s(pace_str):
    """'5:23' → 323 secondes. Retourne None si invalide."""
    if not pace_str: return None
    try:
        parts = str(pace_str).split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except: return None

def s_to_pace_str(secs):
    """323 → '5:23'"""
    if not secs: return "—"
    return f"{int(secs)//60}:{int(secs)%60:02d}"

def get_week_key(date_str):
    """Retourne le lundi de la semaine d'une date."""
    try:
        d = datetime.strptime(date_str[:10], "%Y-%m-%d")
        return (d - timedelta(days=d.weekday())).strftime("%Y-%m-%d")
    except: return date_str[:10]

def get_exercise_date(ex):
    """Extrait la date d'un exercice (multi-format)."""
    for k in ("start_time", "start-time", "date"):
        v = ex.get(k, "")
        if v: return str(v)[:10]
    return ""

def get_exercise_duration_s(ex):
    """Retourne la durée en secondes."""
    if ex.get("duration_s"): return int(ex["duration_s"])
    return int(parse_duration_s(ex.get("duration", "")))

def get_exercise_distance_m(ex):
    """Retourne la distance en mètres."""
    for k in ("distance_m",):
        if ex.get(k): return float(ex[k])
    if ex.get("distance_km"): return float(ex["distance_km"]) * 1000
    if ex.get("distance"):    return float(ex["distance"])
    return 0

def get_splits(ex):
    """Retourne les splits normalisés quel que soit le format."""
    raw = ex.get("splits") or ex.get("_splits_km") or ex.get("_km_splits") or []
    result = []
    for s in raw:
        result.append({
            "km":      s.get("km"),
            "pace":    s.get("pace"),
            "pace_s":  s.get("pace_s") or pace_str_to_s(s.get("pace")),
            "hr_avg":  s.get("hr_avg"),
            "hr_min":  s.get("hr_min"),
            "hr_max":  s.get("hr_max"),
            "cadence": s.get("cadence") or s.get("cadence_spm"),
            "pwr_avg": s.get("pwr_avg") or s.get("power_avg"),
            "pwr_max": s.get("pwr_max") or s.get("power_max"),
            "d_plus":  s.get("d_plus"),
            "d_minus": s.get("d_minus"),
            "d_net":   s.get("d_net"),
        })
    return result

def get_hr_avg(ex):
    """FC moyenne multi-format."""
    if ex.get("hr_avg"): return ex["hr_avg"]
    hr = ex.get("heart_rate") or ex.get("heart-rate") or {}
    if isinstance(hr, dict): return hr.get("average")
    return None

def get_hr_max(ex):
    if ex.get("hr_max"): return ex["hr_max"]
    hr = ex.get("heart_rate") or ex.get("heart-rate") or {}
    if isinstance(hr, dict): return hr.get("maximum")
    return None

def get_cardio_load(ex):
    """Charge cardiaque multi-format."""
    if ex.get("cardio_load"): return float(ex["cardio_load"])
    tl = ex.get("training_load_pro") or ex.get("training-load-pro") or {}
    if isinstance(tl, dict):
        v = tl.get("cardio-load") or tl.get("cardio_load")
        if v: return float(v)
    return 0.0

def get_vo2max(ex):
    return ex.get("vo2max") or ex.get("running_index") or ex.get("running-index")

def assign_zone(hr, zones):
    """Assigne une zone FC à partir d'une valeur de FC."""
    if not hr: return None
    for z in zones:
        if z["min"] <= hr <= z["max"]:
            return z["zone"]
    return None


# ─── CHARGEMENT FICHIERS ───────────────────────────────────────────────────────

def load_json(path, default=None):
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except: return default if default is not None else {}

def load_zones():
    raw = load_json(ZONES_FILE, [])
    # Normalise : ajoute 'zone' si absent (ancien format avec 'lbl')
    result = []
    for z in raw:
        lbl_val = z.get("lbl") or z.get("zone") or z.get("label") or "?"
        result.append({
            "zone":      lbl_val,
            "lbl":       lbl_val,
            "label":     lbl_val,
            "min":       z.get("min", 0),
            "max":       z.get("max", 999),
            "col":       z.get("col", "#ccc"),
            "pace":      z.get("pace", ""),
            "pace_min_s": z.get("pace_min_s", 0),
            "pace_max_s": z.get("pace_max_s", 999),
            "confiance": z.get("confiance", 1.0),
            "updated_at": z.get("updated_at", ""),
        })
    return result

def load_races():
    raw = load_json(RACES_FILE, [])
    if not raw:
        # Fallback défaut
        return [
            {"id":"semi",    "label":"Semi Joinville",  "date":"2026-04-12","objectif":"Sub 1h45-1h50","couleur":"#1565c0"},
            {"id":"marathon","label":"Marathon ING",     "date":"2026-05-16","objectif":"Sub 3h15-3h20","couleur":"#2d6a4f"},
            {"id":"im703",   "label":"IM 70.3 Aix",      "date":"2027-05-16","objectif":"Sub 5h30-6h00","couleur":"#006d77"},
            {"id":"imfull",  "label":"IM Full Vitoria",  "date":"2027-07-11","objectif":"Sub 12h00",    "couleur":"#1a1a2e"},
        ]
    return raw

def load_all_weeks():
    """Charge tous les week_*.json triés du plus ancien au plus récent."""
    weeks = []
    if not WEEKS_DIR.exists(): return weeks
    for f in sorted(WEEKS_DIR.glob("week_2*.json")):
        d = load_json(f, {})
        if d: weeks.append((f.name, d))
    # Ajoute week_current en dernier si pas déjà inclus
    wc = WEEKS_DIR / "week_current.json"
    if wc.exists():
        d = load_json(wc, {})
        if d and d.get("week_key"):
            # Éviter doublon si week_current correspond à une semaine déjà archivée
            existing_keys = {w.get("week_key") for _, w in weeks}
            if d["week_key"] not in existing_keys:
                weeks.append(("week_current.json", d))
    return weeks

def find_programme_file(sem_label):
    """Trouve semaine_XX.json depuis le label 'S12' → 'semaine_12.json'."""
    if not sem_label: return None
    m = re.search(r'(\d+)', str(sem_label))
    if not m: return None
    n = int(m.group(1))
    f = PROG_DIR / f"semaine_{n:02d}.json"
    return f if f.exists() else None

def load_calendar_constraints(week_start, week_end):
    """
    Retourne un dict {date_str: {"type": "voyage"|"depart"|"normal", 
                                  "sport_interdit": [...], "note": str}}
    pour chaque jour de la période week_start → week_end.
    
    Règles couleur :
      #f59e0b = voyage multi-jours → aucun sport
      #3b82f6 = billet/déplacement → course uniquement (vélo+natation interdits)
    """
    events = load_json(CAL_FILE, [])
    voyages = load_json(VOYAGES_FILE, {}).get("voyages", [])

    # Construire le set des jours de voyage depuis voyages.json (confirmation)
    voyage_days = set()
    for v in voyages:
        try:
            d1 = datetime.strptime(v["date_debut"], "%Y-%m-%d")
            d2 = datetime.strptime(v["date_fin"],   "%Y-%m-%d")
            for i in range((d2-d1).days + 1):
                voyage_days.add((d1 + timedelta(days=i)).strftime("%Y-%m-%d"))
        except: pass

    # Indexer les événements calendrier par date
    cal_by_date = defaultdict(list)
    for e in events:
        # Événement multi-jours : étendre sur tous les jours
        try:
            d1 = datetime.strptime(e["date"],     "%Y-%m-%d")
            d2 = datetime.strptime(e["end_date"], "%Y-%m-%d")
            for i in range((d2-d1).days + 1):
                day = (d1 + timedelta(days=i)).strftime("%Y-%m-%d")
                cal_by_date[day].append(e)
        except: pass

    # Construire le résultat pour chaque jour de la semaine cible
    result = {}
    try:
        d_start = datetime.strptime(week_start, "%Y-%m-%d")
        d_end   = datetime.strptime(week_end,   "%Y-%m-%d")
    except: return result

    current = d_start
    while current <= d_end:
        day_str = current.strftime("%Y-%m-%d")
        day_events = cal_by_date.get(day_str, [])

        constraint = {"type": "normal", "sport_interdit": [], "note": ""}

        # Vérifier voyage jaune (#f59e0b)
        voyage_event = next((e for e in day_events if e.get("color") == "#f59e0b"), None)
        if voyage_event or day_str in voyage_days:
            constraint = {
                "type": "voyage",
                "sport_interdit": ["Velo", "Natation", "CYCLING", "SWIMMING"],
                "note": voyage_event["title"] if voyage_event else "Voyage (voyages.json)"
            }
        else:
            # Vérifier départ/retour bleu (#3b82f6)
            depart_event = next((e for e in day_events if e.get("color") == "#3b82f6"), None)
            if depart_event:
                constraint = {
                    "type": "depart",
                    "sport_interdit": ["Velo", "Natation", "CYCLING", "SWIMMING"],
                    "note": depart_event["title"]
                }

        result[day_str] = constraint
        current += timedelta(days=1)

    return result


# ─── LECTURE PROGRAMME EXCEL ─────────────────────────────────────────────────

def load_programme_from_excel():
    """
    Lit Marathon_programme.xlsx directement comme source de vérité.
    Retourne un dict {numéro_semaine: [séances]} avec le format complet
    incluant D+ dans la structure.
    """
    try:
        from openpyxl import load_workbook
        from datetime import datetime as _dt
    except ImportError:
        return {}

    xlsx_path = POLAR / "Marathon_programme.xlsx"
    if not xlsx_path.exists():
        return {}

    try:
        wb = load_workbook(str(xlsx_path), read_only=True)
        ws = wb.active
    except Exception:
        return {}

    programme = {}
    current_sem = None
    current_num = None

    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue

        # Ligne semaine header
        if isinstance(row[0], str) and "Semaine" in row[0]:
            current_sem = row[0]
            m = re.search(r'(\d+)', current_sem)
            current_num = int(m.group(1)) if m else None
            if current_num:
                programme[current_num] = []
            continue

        # Ligne séance
        date_val = row[0]
        if isinstance(date_val, _dt) and current_num is not None:
            seance = {
                "date":        date_val.strftime("%Y-%m-%d"),
                "jour":        row[1] or "",
                "sport":       row[2] or "",
                "type":        row[3] or "",
                "duree":       f"{row[4]}min" if row[4] else "",
                "zone":        row[5] or "",
                "rpe":         str(row[6]) if row[6] else "",
                "structure":   row[7] or "",
                "dist_cible":  float(row[8]) if row[8] and str(row[8]) != "None" else None,
                "dist_velo_km": float(row[9]) if row[9] and str(row[9]) != "None" else None,
                "dist_cible_m": int(row[10]) if row[10] and str(row[10]) != "None" else None,
            }
            programme[current_num].append(seance)

    return programme


# ─── CALCUL SUB CIBLES DEPUIS DONNÉES PHYSIOLOGIQUES ─────────────────────────

def compute_sub_cibles(all_weeks, zones, analysis_data=None):
    """
    Estime les temps par discipline depuis TOUTES les semaines historiques.
    Natation: pace_100m moyen pondéré récence
    Vélo: vitesse moyenne pondérée récence  
    Course: allure Z2 + résultats courses réelles
    Triathlon: combinaison des trois disciplines
    """
    if analysis_data is None:
        analysis_data = {}

    from datetime import datetime, timedelta
    today = datetime.now().date()

    # ── Collecte toutes séances historiques ──────────────────────────────
    run_sessions = []   # {date, dist_km, dur_s, pace_s_km, hr_avg, is_race}
    swim_sessions = []  # {date, dist_m, dur_s, pace_100m_s, hr_avg}
    bike_sessions = []  # {date, dist_km, dur_s, speed_kmh, hr_avg}

    SWIM_SPORTS = {"POOL_SWIMMING", "SWIMMING"}
    BIKE_SPORTS = {"CYCLING", "ROAD_BIKING"}

    for _item in all_weeks:
        week_data = _item[1] if isinstance(_item, tuple) else _item
        for ex in week_data.get("exercises", []):
            sport = ex.get("sport", "").upper()
            date_str = ex.get("date", "")
            try:
                ex_date = datetime.strptime(date_str[:10], "%Y-%m-%d").date()
            except:
                continue
            days_ago = (today - ex_date).days
            # Poids récence : 1.0 si aujourd'hui, 0.3 si >180j
            recency_w = max(0.3, 1.0 - days_ago / 300)

            if sport == "RUNNING":
                dist = ex.get("distance_km") or 0
                dur_s = ex.get("duration_s") or (ex.get("duration_min", 0) * 60)
                if dist > 0 and dur_s > 0:
                    pace_s = dur_s / dist
                    hr = ex.get("hr_avg") or ex.get("heart_rate_avg") or 0
                    is_race = dist >= 19 and dist <= 45 and pace_s < 320  # course rapide
                    run_sessions.append({
                        "date": ex_date, "dist_km": dist, "dur_s": dur_s,
                        "pace_s": pace_s, "hr": hr, "recency": recency_w,
                        "is_race": is_race
                    })

            elif sport in SWIM_SPORTS:
                dist_m = ex.get("distance_m") or 0
                dur_s = ex.get("duration_s") or 0
                pace_100m = ex.get("pace_100m")  # format "M:SS" ou secondes
                if isinstance(pace_100m, str) and ":" in pace_100m:
                    # Format "1:32" ou "1:32/100m"
                    clean = pace_100m.split("/")[0].strip()
                    parts = clean.split(":")
                    try:
                        pace_100m_s = int(parts[0]) * 60 + int(parts[1])
                    except: continue
                elif isinstance(pace_100m, (int, float)):
                    pace_100m_s = float(pace_100m)
                elif dist_m > 0 and dur_s > 0:
                    pace_100m_s = dur_s / dist_m * 100
                else:
                    continue
                if dist_m > 0 and pace_100m_s > 60:  # sanity check
                    swim_sessions.append({
                        "date": ex_date, "dist_m": dist_m, "dur_s": dur_s,
                        "pace_100m_s": pace_100m_s, "recency": recency_w
                    })

            elif sport in BIKE_SPORTS:
                dist = ex.get("distance_km") or 0
                dur_s = ex.get("duration_s") or (ex.get("duration_min", 0) * 60)
                speed = ex.get("speed_avg_kmh") or (dist / (dur_s / 3600) if dur_s > 0 else 0)
                if dist > 0 and speed > 5:
                    bike_sessions.append({
                        "date": ex_date, "dist_km": dist, "dur_s": dur_s,
                        "speed_kmh": speed, "recency": recency_w
                    })

    def s_to_hms(total_s):
        total_s = int(total_s)
        h = total_s // 3600
        m = (total_s % 3600) // 60
        s = total_s % 60
        return f"{h}:{m:02d}:{s:02d}"

    def s_to_pace(pace_s):
        pace_s = int(pace_s)
        return f"{pace_s//60}:{pace_s%60:02d}"

    def weighted_avg(values, weights):
        if not values: return None
        sw = sum(weights)
        if sw == 0: return None
        return sum(v*w for v,w in zip(values, weights)) / sw

    # ── COURSE À PIED ────────────────────────────────────────────────────
    # Trouver meilleure performance récente comme ancre
    best_semi_pace_s = None
    best_semi_date = None
    for s in run_sessions:
        if 19 <= s["dist_km"] <= 23 and s["pace_s"] < 320:  # semi ~4:00-5:20/km
            if best_semi_pace_s is None or s["pace_s"] < best_semi_pace_s:
                best_semi_pace_s = s["pace_s"]
                best_semi_date = s["date"]

    # Allure Z2 pondérée
    z2_sessions = []
    for s in run_sessions:
        hr = s.get("hr", 0)
        pace = s.get("pace_s", 0)
        if hr > 0 and 120 <= hr <= 155 and pace > 240:  # Z2 typique
            z2_sessions.append(s)

    if z2_sessions:
        z2_paces = [s["pace_s"] for s in z2_sessions]
        z2_weights = [s["recency"] for s in z2_sessions]
        z2_pace_avg_s = weighted_avg(z2_paces, z2_weights)
    elif run_sessions:
        all_paces = [s["pace_s"] for s in run_sessions if s["pace_s"] > 0]
        z2_pace_avg_s = sum(all_paces)/len(all_paces) if all_paces else 330
    else:
        z2_pace_avg_s = 330

    # TSB / efficacité bonus
    tsb = analysis_data.get("tsb", 0) or 0
    eff_ratio = analysis_data.get("efficiency_ratio", 1.0) or 1.0
    hr_noc = analysis_data.get("hr_nocturne_moy", 0) or 0
    vol_km = analysis_data.get("volume_km", 0) or 0

    # Pénalité fatigue sur allure (secondes/km)
    fatigue_pen = 0
    if tsb < -20: fatigue_pen += 15
    elif tsb < -10: fatigue_pen += 8
    elif tsb < 0: fatigue_pen += 3
    if eff_ratio > 1.30: fatigue_pen += 10
    elif eff_ratio > 1.15: fatigue_pen += 5
    if hr_noc > 62: fatigue_pen += 5
    elif hr_noc > 57: fatigue_pen += 2

    # Si on a un résultat semi réel, l'utiliser comme ancre
    if best_semi_pace_s:
        days_since_semi = (today - best_semi_date).days if best_semi_date else 999
        # Dégradation si vieux + fatigue actuelle
        semi_pace_s = best_semi_pace_s + fatigue_pen + (days_since_semi // 30) * 2
    else:
        # Extrapolation depuis Z2
        semi_pace_s = z2_pace_avg_s - 30 + fatigue_pen  # semi ~30s/km plus vite que Z2

    # Extrapolation autres distances
    pace_10k_s  = semi_pace_s - 15   # 10km ~15s/km plus vite que semi
    marathon_pace_s = semi_pace_s + 20 + fatigue_pen // 2  # marathon ~20s/km plus lent
    # Formule Riegel : t2 = t1 * (d2/d1)^1.06
    semi_s    = semi_pace_s * 21.097
    marathon_s = marathon_pace_s * 42.195
    pace10k_s = pace_10k_s * 10

    # ── PROBABILITÉS COURSE ──────────────────────────────────────────────
    nb_z2 = len(z2_sessions)
    base_semi     = 65 if nb_z2 >= 3 else 58
    base_marathon = 58 if nb_z2 >= 3 else 48
    tsb_bonus = 8 if tsb > 5 else (5 if tsb > 0 else (-3 if tsb > -10 else (-8 if tsb > -20 else -15)))
    eff_bonus = 8 if eff_ratio < 1.05 else (-3 if eff_ratio < 1.15 else (-8 if eff_ratio < 1.30 else -15))
    vol_bonus = 5 if vol_km >= 50 else (3 if vol_km >= 40 else (-3 if vol_km < 30 else 0))
    hr_bonus  = 0 if hr_noc == 0 else (3 if hr_noc < 52 else (0 if hr_noc < 57 else (-3 if hr_noc < 62 else -8)))
    anchor_bonus = 5 if best_semi_pace_s else 0  # bonus si on a un résultat réel

    proba_10k      = min(88, max(30, base_semi + 5 + tsb_bonus + eff_bonus + vol_bonus + hr_bonus + anchor_bonus))
    proba_semi     = min(88, max(25, base_semi + tsb_bonus + eff_bonus + vol_bonus + hr_bonus + anchor_bonus))
    proba_marathon = min(82, max(20, base_marathon + tsb_bonus + eff_bonus + vol_bonus + hr_bonus + anchor_bonus))

    # ── NATATION ────────────────────────────────────────────────────────
    if swim_sessions:
        paces = [s["pace_100m_s"] for s in swim_sessions]
        weights = [s["recency"] for s in swim_sessions]
        pace_100m_avg_s = weighted_avg(paces, weights)
        nb_swim = len(swim_sessions)
    else:
        pace_100m_avg_s = 115  # 1:55/100m défaut
        nb_swim = 0

    def swim_time(dist_m, pace_100m_s):
        # Pénalité distance : ~3s/100m tous les 1000m
        pen = (dist_m / 1000) * 3
        return dist_m / 100 * (pace_100m_s + pen)

    swim_900_s   = swim_time(900,   pace_100m_avg_s)
    swim_1500_s  = swim_time(1500,  pace_100m_avg_s)
    swim_1900_s  = swim_time(1900,  pace_100m_avg_s)
    swim_3800_s  = swim_time(3800,  pace_100m_avg_s)
    proba_swim = min(80, max(30, 55 + (nb_swim * 3) + tsb_bonus // 2))

    # ── VÉLO ─────────────────────────────────────────────────────────────
    if bike_sessions:
        speeds = [s["speed_kmh"] for s in bike_sessions]
        weights = [s["recency"] for s in bike_sessions]
        speed_avg = weighted_avg(speeds, weights)
        nb_bike = len(bike_sessions)
    else:
        speed_avg = 28.0  # défaut
        nb_bike = 0

    def bike_time(dist_km, speed_kmh):
        # Pénalité distance : vitesse baisse légèrement sur longue durée
        pen_factor = 1.0 + (dist_km / 200) * 0.05
        return dist_km / (speed_kmh / pen_factor) * 3600

    bike_40_s  = bike_time(40,  speed_avg)
    bike_90_s  = bike_time(90,  speed_avg)
    bike_180_s = bike_time(180, speed_avg)
    proba_bike = min(80, max(30, 50 + (nb_bike * 5) + tsb_bonus // 2))

    # ── TRIATHLON ────────────────────────────────────────────────────────
    # T1/T2 transitions
    T1 = 3 * 60  # 3min
    T2 = 2 * 60  # 2min

    im703_s  = swim_1900_s + T1 + bike_90_s  + T2 + semi_s
    imfull_s = swim_3800_s + T1 + bike_180_s + T2 + marathon_s
    proba_im703  = min(78, max(20, (proba_swim + proba_bike + proba_semi) // 3))
    proba_imfull = min(72, max(15, (proba_swim + proba_bike + proba_marathon) // 3 - 5))

    data_quality = []
    if nb_swim == 0: data_quality.append("natation: pas de données")
    if nb_bike == 0: data_quality.append("vélo: peu de données")
    if not best_semi_pace_s: data_quality.append("course: pas de résultat race")

    return {
        "z2_pace_avg":        s_to_pace(z2_pace_avg_s) if z2_pace_avg_s else "?",
        "z2_pace_avg_s":      z2_pace_avg_s,
        "best_efficiency":    analysis_data.get("efficacite_best"),
        "current_efficiency": analysis_data.get("efficacite_moy"),
        "efficiency_ratio":   eff_ratio,
        "nb_z2_sessions":     nb_z2,
        "data_quality":       data_quality,
        # Course
        "run_10k":   {"sub_cible": s_to_hms(pace10k_s),  "allure": s_to_pace(pace_10k_s),  "probabilite_pct": proba_10k,      "base": f"Ancre semi {s_to_pace(best_semi_pace_s)}/km" if best_semi_pace_s else "Extrapolation Z2"},
        "semi":      {"sub_cible": s_to_hms(semi_s),     "allure": s_to_pace(semi_pace_s),  "probabilite_pct": proba_semi,     "base": f"Résultat {best_semi_date}" if best_semi_pace_s else "Extrapolation Z2", "sub_cible_s": int(semi_s), "allure_cible": s_to_pace(semi_pace_s)},
        "marathon":  {"sub_cible": s_to_hms(marathon_s), "allure": s_to_pace(marathon_pace_s), "probabilite_pct": proba_marathon, "base": "Extrapolation semi +Riegel", "sub_cible_s": int(marathon_s), "allure_cible": s_to_pace(marathon_pace_s)},
        # Natation
        "swim_900m":  {"sub_cible": s_to_hms(swim_900_s),  "pace_100m": s_to_pace(pace_100m_avg_s), "probabilite_pct": proba_swim, "nb_sessions": nb_swim},
        "swim_1500m": {"sub_cible": s_to_hms(swim_1500_s), "pace_100m": s_to_pace(pace_100m_avg_s), "probabilite_pct": proba_swim, "nb_sessions": nb_swim},
        "swim_1900m": {"sub_cible": s_to_hms(swim_1900_s), "pace_100m": s_to_pace(pace_100m_avg_s), "probabilite_pct": proba_swim, "nb_sessions": nb_swim},
        "swim_3800m": {"sub_cible": s_to_hms(swim_3800_s), "pace_100m": s_to_pace(pace_100m_avg_s), "probabilite_pct": proba_swim, "nb_sessions": nb_swim},
        # Vélo
        "bike_40km":  {"sub_cible": s_to_hms(bike_40_s),  "speed": round(speed_avg, 1), "probabilite_pct": proba_bike, "nb_sessions": nb_bike},
        "bike_90km":  {"sub_cible": s_to_hms(bike_90_s),  "speed": round(speed_avg, 1), "probabilite_pct": proba_bike, "nb_sessions": nb_bike},
        "bike_180km": {"sub_cible": s_to_hms(bike_180_s), "speed": round(speed_avg, 1), "probabilite_pct": proba_bike, "nb_sessions": nb_bike},
        # Triathlon
        "im703":  {"sub_cible": s_to_hms(im703_s),  "probabilite_pct": proba_im703,  "base": f"Nat {s_to_hms(swim_1900_s)} + Vélo {s_to_hms(bike_90_s)} + Run {s_to_hms(semi_s)}"},
        "imfull": {"sub_cible": s_to_hms(imfull_s), "probabilite_pct": proba_imfull, "base": f"Nat {s_to_hms(swim_3800_s)} + Vélo {s_to_hms(bike_180_s)} + Run {s_to_hms(marathon_s)}"},
        # Compat ancien format
        "im703_compat":  {"sub_cible": s_to_hms(im703_s),  "allure_cible": None, "probabilite_pct": proba_im703,  "base": "Estimation combinée"},
        "imfull_compat": {"sub_cible": s_to_hms(imfull_s), "allure_cible": None, "probabilite_pct": proba_imfull, "base": "Estimation combinée"},
    }

def compute_atl_ctl_tsb(all_weeks):
    """
    Calcule ATL (7j), CTL (42j), TSB = CTL - ATL sur tout l'historique.
    Inclut tous les sports (ski, vélo, course, natation).
    Retourne le dernier état + historique.
    """
    # Agréger la charge par date
    daily_load = defaultdict(float)
    for _, wd in all_weeks:
        for ex in wd.get("exercises", []):
            date = get_exercise_date(ex)
            load = get_cardio_load(ex)
            if date and load:
                daily_load[date] += load

    if not daily_load:
        return {"atl": 0, "ctl": 0, "tsb": 0, "history": []}

    dates = sorted(daily_load.keys())
    start = datetime.strptime(dates[0], "%Y-%m-%d")
    end   = datetime.now()
    all_dates = [(start + timedelta(days=i)).strftime("%Y-%m-%d")
                 for i in range((end - start).days + 2)]

    atl = 0.0; ctl = 0.0
    k_atl = 1 - 1/7; k_ctl = 1 - 1/42
    history = []

    for dt in all_dates:
        load = daily_load.get(dt, 0.0)
        atl  = atl * k_atl + load * (1 - k_atl)
        ctl  = ctl * k_ctl + load * (1 - k_ctl)
        tsb  = ctl - atl
        if load > 0:
            history.append({"date": dt, "load": round(load, 1),
                            "atl": round(atl, 1), "ctl": round(ctl, 1),
                            "tsb": round(tsb, 1)})

    return {
        "atl":     round(atl, 1),
        "ctl":     round(ctl, 1),
        "tsb":     round(ctl - atl, 1),
        "history": history[-14:]  # 14 derniers jours avec charge
    }


# ─── ANALYSE SÉANCES ──────────────────────────────────────────────────────────

def analyse_exercise(ex, zones):
    """Enrichit un exercice avec les métriques calculées."""
    sport    = ex.get("sport", "?")
    date     = get_exercise_date(ex)
    dur_s    = get_exercise_duration_s(ex)
    dist_m   = get_exercise_distance_m(ex)
    hr_avg   = get_hr_avg(ex)
    hr_max   = get_hr_max(ex)
    load     = get_cardio_load(ex)
    vo2      = get_vo2max(ex)
    splits   = get_splits(ex)

    # ── Calculer les champs manquants depuis les splits (Format C) ──
    # pace_avg : moyenne des pace_s des splits complets (pas le dernier partiel)
    pace_avg_s = pace_str_to_s(ex.get("pace_avg"))
    if not pace_avg_s and splits:
        full_splits = [s for s in splits if s.get("pace_s") and
                       not str(s.get("km","")).endswith("+")]
        if full_splits:
            pace_avg_s = round(sum(s["pace_s"] for s in full_splits) / len(full_splits))

    # pace_best : km le plus rapide
    if not ex.get("pace_best") and splits:
        best_s = min((s["pace_s"] for s in splits if s.get("pace_s") and s["pace_s"] > 0),
                     default=None)
        if best_s:
            ex = dict(ex)
            ex["pace_best"] = s_to_pace_str(best_s)

    # cadence_avg : moyenne des cadences splits
    cadence_avg = ex.get("cadence_avg")
    if not cadence_avg and splits:
        cads = [s["cadence"] for s in splits if s.get("cadence")]
        if cads:
            cadence_avg = round(sum(cads) / len(cads))

    # power_avg : moyenne des puissances splits
    power_avg = ex.get("power_avg")
    if not power_avg and splits:
        pwrs = [s["pwr_avg"] for s in splits if s.get("pwr_avg")]
        if pwrs:
            power_avg = round(sum(pwrs) / len(pwrs))

    # distance_m : fallback sur champ 'distance' (Format C)
    if not dist_m:
        dist_m = get_exercise_distance_m(ex)

    # Efficacité aérobie running (pace_s / fc_avg — plus bas = meilleur)
    aerobic_efficiency = None
    if sport in ("RUNNING",) and pace_avg_s and hr_avg:
        aerobic_efficiency = round(pace_avg_s / hr_avg, 3)

    # Répartition splits par zone
    zone_dist = {"Z1": 0, "Z2": 0, "Z3": 0, "Z4": 0, "Z5": 0, "?": 0}
    for s in splits:
        z = assign_zone(s.get("hr_avg"), zones)
        zone_dist[z or "?"] += 1

    total_splits = sum(zone_dist.values())
    zone_pct = {}
    if total_splits > 0:
        zone_pct = {z: round(c/total_splits*100) for z, c in zone_dist.items()}

    # D+ total depuis splits
    d_plus_total = sum(s.get("d_plus") or 0 for s in splits)

    # Mix énergétique
    carb_pct = ex.get("kcal_carb_pct") or ex.get("carbohydrate_percentage") or \
               ex.get("carbohydrate-percentage")
    fat_pct  = ex.get("kcal_fat_pct")  or ex.get("fat_percentage") or \
               ex.get("fat-percentage")

    # Ratio muscle/cardio
    muscle_load  = ex.get("muscle_load")
    muscle_level = ex.get("muscle_load_level")
    cardio_level = ex.get("cardio_load_level")
    ratio_mc = None
    if muscle_load and load and load > 0:
        ratio_mc = round(float(muscle_load) / load, 2)

    # pace_best
    pace_best_s = pace_str_to_s(ex.get("pace_best"))

    return {
        "date":               date,
        "sport":              sport,
        "duration_s":         dur_s,
        "duration_fmt":       ex.get("duration", s_to_pace_str(dur_s)),
        "distance_km":        round(dist_m / 1000, 2) if dist_m else None,
        "pace_avg":           ex.get("pace_avg") or (s_to_pace_str(pace_avg_s) if pace_avg_s else None),
        "pace_avg_s":         pace_avg_s,
        "pace_best":          ex.get("pace_best"),
        "pace_best_s":        pace_best_s,
        "pace_100m":          ex.get("pace_100m"),
        "speed_avg_kmh":      ex.get("speed_avg_kmh"),
        "hr_avg":             hr_avg,
        "hr_max":             hr_max,
        "cardio_load":        round(load, 1) if load else None,
        "cardio_load_level":  cardio_level,
        "muscle_load":        round(float(muscle_load), 1) if muscle_load else None,
        "muscle_load_level":  muscle_level,
        "ratio_muscle_cardio": ratio_mc,
        "vo2max":             vo2,
        "power_avg":          power_avg,
        "cadence_avg":        cadence_avg,
        "elevation_up_m":     ex.get("elevation_up_m") or (round(d_plus_total, 1) if d_plus_total > 0 else None),
        "elevation_down_m":   ex.get("elevation_down_m"),
        "kcal":               ex.get("kcal") or ex.get("calories"),
        "carb_pct":           carb_pct,
        "fat_pct":            fat_pct,
        "aerobic_efficiency": aerobic_efficiency,
        "zone_distribution":  zone_pct,
        "splits":             splits,
        "nb_splits":          len(splits),
    }

def summarise_week_exercises(exercises_analysed):
    """Résumé agrégé d'une semaine d'exercices."""
    by_sport = defaultdict(lambda: {"km": 0, "h": 0, "load": 0, "nb": 0})
    all_load = 0

    for ex in exercises_analysed:
        sp = ex["sport"]
        by_sport[sp]["km"]   += ex.get("distance_km") or 0
        by_sport[sp]["h"]    += (ex.get("duration_s") or 0) / 3600
        by_sport[sp]["load"] += ex.get("cardio_load") or 0
        by_sport[sp]["nb"]   += 1
        all_load += ex.get("cardio_load") or 0

    return {
        "par_sport":   {sp: {k: round(v, 2) for k, v in d.items()}
                        for sp, d in by_sport.items()},
        "charge_totale": round(all_load, 1),
    }


# ─── SOMMEIL ──────────────────────────────────────────────────────────────────

def analyse_sleep(sleep_list):
    """Calcule les métriques sommeil de la semaine."""
    if not sleep_list:
        return {}

    scores  = [s.get("sleep_score") or s.get("score")     for s in sleep_list if s.get("sleep_score") or s.get("score")]
    totals  = [(s.get("total_sleep_minutes") or s.get("total_min")) for s in sleep_list if s.get("total_sleep_minutes") or s.get("total_min")]
    hr_nocs = [s.get("heart_rate_avg") or s.get("hr_avg") for s in sleep_list if s.get("heart_rate_avg") or s.get("hr_avg")]
    deeps   = [(s.get("deep_sleep",0)//60 if s.get("deep_sleep") else None) or s.get("phases", {}).get("deep_min")  for s in sleep_list
               if s.get("phases", {}).get("deep_min")]
    rems    = [(s.get("rem_sleep",0)//60 if s.get("rem_sleep") else None) or s.get("phases", {}).get("rem_min")   for s in sleep_list
               if s.get("phases", {}).get("rem_min")]
    # recharge_status = string ("DEPLETED","COMPROMISED","SUSTAINED","EXCELLENT")
    _recharge_map = {"EXCELLENT":5,"SUSTAINED":4,"COMPROMISED":3,"DEPLETED":2,"UNKNOWN":1}
    recharges = [_recharge_map.get(s.get("nightly_recharge_status") or s.get("recharge_status",""),None)
                 for s in sleep_list if s.get("nightly_recharge_status") or s.get("recharge_status")]
    recharges = [r for r in recharges if r is not None]

    # Trend FC nocturne nuit par nuit
    hr_trend = [{"date": s.get("date"), "hr_avg": s.get("hr_avg"),
                 "score": s.get("score"), "recharge": s.get("recharge_status")}
                for s in sleep_list]

    return {
        "nb_nuits":          len(sleep_list),
        "score_moyen":       round(sum(scores)/len(scores), 1)  if scores  else None,
        "score_min":         min(scores)                         if scores  else None,
        "score_max":         max(scores)                         if scores  else None,
        "duree_moy_min":     round(sum(totals)/len(totals), 0)  if totals  else None,
        "fc_nocturne_moy":   round(sum(hr_nocs)/len(hr_nocs), 1) if hr_nocs else None,
        "fc_nocturne_min":   min(hr_nocs)                         if hr_nocs else None,
        "fc_nocturne_max":   max(hr_nocs)                         if hr_nocs else None,
        "fc_trend_7j":       hr_trend,
        "deep_moy_min":      round(sum(deeps)/len(deeps), 0)    if deeps   else None,
        "rem_moy_min":       round(sum(rems)/len(rems), 0)      if rems    else None,
        "recharge_moyen":    round(sum(recharges)/len(recharges), 1) if recharges else None,
    }

def correlate_sleep_perf(sleep_list, exercises_analysed):
    """
    Pour chaque séance, retourne le score sommeil de la nuit précédente.
    Permet au coach de détecter la corrélation sommeil J-1 / perf J.
    """
    sleep_by_date = {s["date"]: s for s in sleep_list if s.get("date")}
    result = []
    for ex in exercises_analysed:
        if ex.get("aerobic_efficiency") is None: continue
        date = ex["date"]
        try:
            prev = (datetime.strptime(date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
            sleep = sleep_by_date.get(prev, {})
            result.append({
                "date":              date,
                "sport":             ex["sport"],
                "aerobic_efficiency": ex["aerobic_efficiency"],
                "sleep_score_j1":    sleep.get("sleep_score") or sleep.get("score"),
                "sleep_min_j1":      sleep.get("total_sleep_minutes") or sleep.get("total_min"),
                "hr_nocturne_j1":    sleep.get("heart_rate_avg") or sleep.get("hr_avg"),
            })
        except: pass
    return result


# ─── PROGRESSION 4 SEMAINES ───────────────────────────────────────────────────

def compute_trends(all_weeks, zones, current_week_key):
    """
    Calcule les tendances sur les 4 semaines précédant la semaine analysée.
    """
    # Trouver les 4 dernières semaines avant current_week_key
    week_keys_sorted = sorted(
        [(wd.get("week_key", ""), wd) for _, wd in all_weeks],
        key=lambda x: x[0]
    )
    prior_weeks = [(wk, wd) for wk, wd in week_keys_sorted if wk < current_week_key][-4:]

    trends = []
    for wk, wd in prior_weeks:
        exs = [analyse_exercise(ex, zones) for ex in wd.get("exercises", [])]
        run_exs = [e for e in exs if e["sport"] == "RUNNING"]

        vo2_vals = [e["vo2max"] for e in run_exs if e.get("vo2max")]
        effic_vals = [e["aerobic_efficiency"] for e in run_exs if e.get("aerobic_efficiency")]
        load_vals  = [e["cardio_load"] for e in exs if e.get("cardio_load")]
        sleep_data = analyse_sleep(wd.get("sleep", []))

        # Volume par sport
        run_km   = sum(e["distance_km"] or 0 for e in exs if e["sport"] == "RUNNING")
        swim_m   = sum((e["distance_km"] or 0)*1000 for e in exs
                       if e["sport"] in ("SWIMMING", "POOL_SWIMMING"))
        bike_km  = sum(e["distance_km"] or 0 for e in exs if e["sport"] == "CYCLING")

        trends.append({
            "week_key":          wk,
            "sem_label":         wd.get("sem_label", ""),
            "run_km":            round(run_km, 1),
            "swim_m":            round(swim_m),
            "bike_km":           round(bike_km, 1),
            "charge_totale":     round(sum(load_vals), 1),
            "vo2max_moyen":      round(sum(vo2_vals)/len(vo2_vals), 1) if vo2_vals else None,
            "efficacite_moy":    round(sum(effic_vals)/len(effic_vals), 3) if effic_vals else None,
            "sleep_score_moyen": (sleep_data or {}).get("score_moyen"),
            "fc_nocturne_moy":   sleep_data.get("fc_nocturne_moy"),
        })
    return trends


# ─── DELTA PRÉVU / RÉALISÉ ────────────────────────────────────────────────────

def compute_delta(programme_seances, exercises_analysed):
    """
    Compare programme prévu vs séances réalisées.
    Match par sport + semaine (pas date exacte).
    """
    realise_by_sport = defaultdict(list)
    for ex in exercises_analysed:
        realise_by_sport[ex["sport"]].append(ex)

    # Mapper sport programme → sport polar
    sport_map = {"Course": "RUNNING", "Velo": "CYCLING",
                 "Natation": "SWIMMING", "POOL_SWIMMING": "SWIMMING"}

    deltas = []
    for s in programme_seances:
        sport_prog  = s.get("sport", "")
        sport_polar = sport_map.get(sport_prog, sport_prog)
        dist_cible  = s.get("dist_cible") or (
            (s.get("dist_cible_m") or 0) / 1000 if s.get("dist_cible_m") else None
        )

        # Trouver la séance réalisée la plus proche
        candidats = realise_by_sport.get(sport_polar, []) + \
                    realise_by_sport.get(sport_prog, [])

        realise = None
        if candidats:
            # Prendre la séance avec la date la plus proche de la date programme
            try:
                prog_dt = datetime.strptime(s["date"], "%Y-%m-%d")
                candidats_sorted = sorted(
                    candidats,
                    key=lambda e: abs((datetime.strptime(e["date"], "%Y-%m-%d") - prog_dt).days)
                )
                # Accepter si dans les 5 jours
                if abs((datetime.strptime(candidats_sorted[0]["date"], "%Y-%m-%d") - prog_dt).days) <= 5:
                    realise = candidats_sorted[0]
            except: pass

        delta = {
            "date_prev":   s["date"],
            "sport":       sport_prog,
            "type":        s.get("type", ""),
            "zone_cible":  s.get("zone", ""),
            "rpe_cible":   s.get("rpe"),
            "dist_cible":  dist_cible,
            "duree_cible": s.get("duree"),
            "structure":   s.get("structure", ""),
            "realise":     None,
            "statut":      "non_realise",
        }

        if realise:
            delta["realise"] = {
                "date":        realise["date"],
                "distance_km": realise["distance_km"],
                "duration_fmt": realise["duration_fmt"],
                "pace_avg":    realise["pace_avg"],
                "hr_avg":      realise["hr_avg"],
                "cardio_load": realise["cardio_load"],
                "elevation_up_m": realise["elevation_up_m"],
                "zone_dominante": max(
                    realise.get("zone_distribution", {"?": 100}),
                    key=lambda z: realise.get("zone_distribution", {}).get(z, 0)
                ) if realise.get("zone_distribution") else None,
            }
            # Déterminer le statut
            if dist_cible and realise["distance_km"]:
                pct = realise["distance_km"] / dist_cible
                if pct >= 0.90:   delta["statut"] = "realise"
                elif pct >= 0.70: delta["statut"] = "partiel"
                else:             delta["statut"] = "sous_realise"
            else:
                delta["statut"] = "realise"

        deltas.append(delta)

    return deltas


# ─── RECALIBRAGE ZONES (Option C) ─────────────────────────────────────────────

def check_zone_recalibration(all_weeks, zones):
    """
    Calibre TOUTES les zones FC+allure depuis les splits réels.
    Pour chaque zone FC, collecte les allures mesurées et calcule min-max.
    """
    from collections import defaultdict
    zone_paces = defaultdict(list)

    for _, wd in all_weeks[-12:]:
        for ex in wd.get("exercises", []):
            if ex.get("sport") != "RUNNING": continue
            splits = get_splits(ex)
            if len(splits) < 4: continue
            for sp in splits:
                hr = sp.get("hr_avg") or sp.get("avg_heart_rate")
                pace = sp.get("pace_s")
                if not hr or not pace or pace <= 0 or pace > 600: continue
                for z in zones:
                    if z["min"] <= hr <= z["max"]:
                        zone_paces[z.get("lbl") or z.get("zone","?")].append(pace)
                        break

    if not any(zone_paces.values()):
        return {"auto_update": False, "pending": True,
                "pending_reason": "Pas assez de données splits"}

    def pct(lst, p):
        lst_s = sorted(lst)
        return lst_s[max(0, int(len(lst_s)*p/100))]

    def fmt_pace(s):
        s = int(s)
        return f"{s//60}:{s%60:02d}"

    proposed_zones = []
    calibration_info = {}
    # Calculer médiane par zone
    zone_medians = {}
    for z in zones:
        lbl = z.get('lbl') or z.get('zone','?')
        paces = zone_paces.get(lbl, [])
        if len(paces) >= 5:
            zone_medians[lbl] = pct(paces, 50)

    # Calculer les bornes strictement monotones
    # Z1(plus lent) → Z5(plus rapide) : pace_max décroissant
    prev_pace_min = 999  # borne inf de la zone précédente (plus lente)
    for idx, z in enumerate(zones):
        lbl = z.get("lbl") or z.get("zone","?")
        paces = zone_paces.get(lbl, [])
        new_z = dict(z)
        if len(paces) >= 5:
            median = pct(paces, 50)
            spread = max(20, 45 - idx * 5)  # spread décroissant : Z1=45s, Z5=25s
            pace_min_s = int(median - spread)
            pace_max_s = int(median + spread)
            # Enforcer monotonie stricte
            if pace_max_s >= prev_pace_min:
                pace_max_s = prev_pace_min - 5
            min_spread = 25
            if pace_max_s - pace_min_s < min_spread:
                pace_min_s = pace_max_s - min_spread
            # Limites réalistes
            pace_min_s = max(220, pace_min_s)  # pas plus rapide que 3:40/km
            pace_max_s = min(480, pace_max_s)  # pas plus lent que 8:00/km
            prev_pace_min = pace_min_s
            if lbl == "Z1":
                new_z["pace"] = f">{fmt_pace(pace_max_s)}"
                new_z["pace_min_s"] = pace_max_s
                new_z["pace_max_s"] = 999
            elif lbl == "Z5":
                new_z["pace"] = f"<{fmt_pace(pace_min_s)}"
                new_z["pace_min_s"] = 0
                new_z["pace_max_s"] = pace_min_s
            else:
                new_z["pace"] = f"{fmt_pace(pace_min_s)}-{fmt_pace(pace_max_s)}"
                new_z["pace_min_s"] = pace_min_s
                new_z["pace_max_s"] = pace_max_s
            new_z["updated_at"] = datetime.now().strftime("%Y-%m-%d")
            calibration_info[lbl] = {"n": len(paces), "median": fmt_pace(int(median))}
        else:
            calibration_info[lbl] = {"n": len(paces), "kept": True}
        proposed_zones.append(new_z)

    z2_ok = len(zone_paces.get("Z2", [])) >= 5
    z3_ok = len(zone_paces.get("Z3", [])) >= 5
    auto_update = False  # Zones gérées manuellement

    return {
        "auto_update": auto_update,
        "proposed_zones": proposed_zones,
        "calibration_info": calibration_info,
        "reason": "Calibration splits: " + ", ".join(f"{z}={calibration_info[z]['n']}" for z in calibration_info),
        "pending": not auto_update,
        "pending_reason": "" if auto_update else "Z2/Z3 insuffisantes",
    }

def build_prompt(
    week_data,        # semaine analysée
    exercises_an,     # exercices analysés
    sleep_data,       # données sommeil
    sleep_corr,       # corrélation sommeil/perf
    atl_ctl,          # métriques fitness
    trends,           # tendances 4 semaines
    delta,            # prévu/réalisé
    prog_current,     # programme semaine analysée
    prog_next,        # programme S+1
    prog_next2,       # programme S+2
    zones,            # zones actuelles
    races,            # courses + objectifs
    constraints,      # contraintes calendrier S+1 et S+2
    context_manuel,   # texte libre utilisateur
    zone_calib,       # résultat check recalibration
    sub_cibles,       # sub cibles pré-calculées
    cal_s1="",        # calendrier exact S+1 (Lundi=date, Mardi=date, ...)
    cal_s2="",        # calendrier exact S+2
):
    today    = datetime.now().strftime("%Y-%m-%d")
    sem_label = week_data.get("sem_label", "?")
    sem_next  = prog_next.get("semaine", "S?") if prog_next else "S?"

    # ── Détection course/compétition dans la semaine analysée ──
    # Une course = séance avec vo2max élevé (≥60), FC très haute (≥170 moy), ou type "race"
    has_race_this_week = False
    race_date_str = ""
    for ex in exercises_an:
        vo2 = ex.get("vo2max") or 0
        hr  = ex.get("hr_avg") or 0
        try: vo2 = float(vo2)
        except: vo2 = 0
        try: hr = float(hr)
        except: hr = 0
        # Détection : VO2max ≥ 60 (indice de performance élevée) ET FC moy ≥ 170
        if vo2 >= 60 and hr >= 170:
            has_race_this_week = True
            race_date_str = ex.get("date", "")
            break
        # Détection secondaire : séance très courte (<35min) avec allure très rapide ET FC élevée
        dur = ex.get("duration_s") or 0
        pace_s = ex.get("pace_avg_s") or 0
        if hr >= 175 and pace_s and pace_s < 270:  # < 4:30/km avec FC ≥ 175
            has_race_this_week = True
            race_date_str = ex.get("date", "")
            break

    # ── Calcul J-X courses ──
    races_jx = []
    for r in races:
        try:
            dt = datetime.strptime(r["date"], "%Y-%m-%d")
            jx = (dt.date() - datetime.strptime(today, "%Y-%m-%d").date()).days
            if jx > 0:
                races_jx.append({
                    "id":       r["id"],
                    "label":    r["label"],
                    "date":     r["date"],
                    "jx":       jx,
                    "objectif": r.get("objectif", ""),
                    "sub_actuel": r.get("objectif", ""),
                })
        except: pass

    # ── Zones formatées ──
    zones_fmt = "\n".join(
        f"  {z['zone']} : FC {z['min']}-{z['max']} bpm | allure {z.get('pace','?')}"
        for z in zones
    )

    # ── Tendances formatées ──
    trends_fmt = ""
    for t in trends:
        trends_fmt += (
            f"  {t['sem_label']} ({t['week_key']}) : "
            f"run={t['run_km']}km, swim={t['swim_m']}m, bike={t['bike_km']}km | "
            f"charge={t['charge_totale']} | vo2={t['vo2max_moyen']} | "
            f"effic={t['efficacite_moy']} | sommeil={t['sleep_score_moyen']} | "
            f"FC_noc={t['fc_nocturne_moy']}\n"
        )

    # ── Séances semaine analysée ──
    exs_fmt = ""
    for ex in exercises_an:
        sp = ex["sport"]
        exs_fmt += f"\n  [{sp}] {ex['date']}"
        exs_fmt += f"\n    Distance: {ex.get('distance_km')}km"
        exs_fmt += f"\n    Durée: {ex.get('duration_fmt')} | Allure moy: {ex.get('pace_avg')} | Meilleur km: {ex.get('pace_best')}"
        if sp == "RUNNING":
            exs_fmt += f"\n    FC: moy={ex.get('hr_avg')} max={ex.get('hr_max')}"
            exs_fmt += f"\n    Puissance moy: {ex.get('power_avg')}W | Cadence: {ex.get('cadence_avg')}spm"
            exs_fmt += f"\n    D+: {ex.get('elevation_up_m')}m"
            exs_fmt += f"\n    Charge cardio: {ex.get('cardio_load')} ({ex.get('cardio_load_level')}) | Charge muscu: {ex.get('muscle_load')} ({ex.get('muscle_load_level')}) | Ratio M/C: {ex.get('ratio_muscle_cardio')}"
            exs_fmt += f"\n    Efficacité aérobie: {ex.get('aerobic_efficiency')} (lower=better)"
            exs_fmt += f"\n    Mix énergétique: {ex.get('carb_pct')}% carb / {ex.get('fat_pct')}% lipides"
            exs_fmt += f"\n    Zones: {ex.get('zone_distribution')}"
            if ex.get("splits"):
                exs_fmt += f"\n    Splits ({ex['nb_splits']} km) :"
                for s in ex["splits"]:
                    exs_fmt += (f"\n      km{s['km']}: {s['pace']} | "
                                f"FC {s['hr_avg']}({s['hr_min']}-{s['hr_max']}) | "
                                f"cad={s['cadence']} | pwr={s['pwr_avg']}W | "
                                f"D+={s['d_plus']}m D-={s['d_minus']}m")
        elif sp in ("SWIMMING", "POOL_SWIMMING"):
            exs_fmt += f"\n    Allure: {ex.get('pace_100m')} | FC: moy={ex.get('hr_avg')} max={ex.get('hr_max')}"
            exs_fmt += f"\n    Charge cardio: {ex.get('cardio_load')}"
        elif sp == "CYCLING":
            exs_fmt += f"\n    Vitesse moy: {ex.get('speed_avg_kmh')} km/h"
            exs_fmt += f"\n    FC: moy={ex.get('hr_avg')} | Puissance: {ex.get('power_avg')}W"
            exs_fmt += f"\n    Charge cardio: {ex.get('cardio_load')}"
        elif "SKIING" in sp:
            exs_fmt += f"\n    Charge cardio: {ex.get('cardio_load')} ({ex.get('cardio_load_level')})"
        exs_fmt += "\n"

    # ── Sommeil ──
    sl = sleep_data
    sleep_fmt = f"""  Score moyen: {sl.get('score_moyen')} (min={sl.get('score_min')}, max={sl.get('score_max')})
  Durée moy: {sl.get('duree_moy_min')} min
  FC nocturne moy: {sl.get('fc_nocturne_moy')} bpm (min={sl.get('fc_nocturne_min')}, max={sl.get('fc_nocturne_max')})
  Deep moy: {sl.get('deep_moy_min')} min | REM moy: {sl.get('rem_moy_min')} min
  Recharge Polar moyen: {sl.get('recharge_moyen')}/5
  Tendance FC nocturne (nuit par nuit):"""
    for n in (sl.get("fc_trend_7j") or []):
        sleep_fmt += f"\n    {n['date']}: score={n['score']} FC={n['hr_avg']} recharge={n['recharge']}"

    # ── Corrélation sommeil/perf ──
    corr_fmt = ""
    for c in sleep_corr:
        corr_fmt += (f"  {c['date']} effic={c['aerobic_efficiency']} | "
                     f"sommeil_J-1: score={c['sleep_score_j1']} min={c['sleep_min_j1']} "
                     f"FC={c['hr_nocturne_j1']}\n")

    # ── Delta prévu/réalisé ──
    delta_fmt = ""
    for d in delta:
        statut_ico = {"realise":"✅","partiel":"⚠️","sous_realise":"❌","non_realise":"🚫"}.get(d["statut"],"?")
        delta_fmt += f"  {statut_ico} {d['date_prev']} [{d['sport']}] {d['type']}\n"
        delta_fmt += f"    Prévu: {d['dist_cible']}km / {d['duree_cible']} / {d['zone_cible']} / RPE {d['rpe_cible']}\n"
        if d.get("realise"):
            r = d["realise"]
            delta_fmt += (f"    Réalisé: {r['distance_km']}km / {r['duration_fmt']} / "
                         f"FC={r['hr_avg']} / load={r['cardio_load']} / "
                         f"D+={r['elevation_up_m']}m / zone={r['zone_dominante']}\n")
        delta_fmt += f"    Structure prévue: {d['structure']}\n\n"

    # ── Programme S+1 ──
    prog_next_fmt = ""
    if prog_next:
        for s in prog_next.get("seances", []):
            c = constraints.get(s.get("date", ""), {})
            contrainte_note = ""
            if c.get("type") != "normal":
                contrainte_note = f" ⚠️ CONTRAINTE: {c['type'].upper()} — {c['note']} → sports interdits: {c['sport_interdit']}"
            # Distance selon sport
            if s.get("dist_cible_m"):
                dist_str = f"dist_cible_m={s['dist_cible_m']} (garder exactement)"
            elif s.get("dist_cible"):
                dist_str = f"dist_cible={s['dist_cible']}km (garder exactement)"
            else:
                dist_str = "dist_cible=null"
            prog_next_fmt += (
                f"  {s['date']} {s['jour']} | {s['sport']} | {s['type']}\n"
                f"    Durée: {s.get('duree','?')} | Zone: {s.get('zone','?')} | "
                f"RPE: {s.get('rpe','?')} | {dist_str}\n"
                f"    Structure: {s.get('structure','?')}\n"
                f"{contrainte_note}\n\n"
            )

    # ── Programme S+2 ──
    prog_next2_fmt = ""
    if prog_next2:
        for s in prog_next2.get("seances", []):
            c = constraints.get(s.get("date", ""), {})
            contrainte_note = ""
            if c.get("type") != "normal":
                contrainte_note = f" ⚠️ CONTRAINTE: {c['type'].upper()} — {c['note']} → sports interdits: {c['sport_interdit']}"
            if s.get("dist_cible_m"):
                dist_str = f"dist_cible_m={s['dist_cible_m']} (garder exactement)"
            elif s.get("dist_cible"):
                dist_str = f"dist_cible={s['dist_cible']}km (garder exactement)"
            else:
                dist_str = "dist_cible=null"
            prog_next2_fmt += (
                f"  {s['date']} {s['jour']} | {s['sport']} | {s['type']}\n"
                f"    Durée: {s.get('duree','?')} | Zone: {s.get('zone','?')} | "
                f"RPE: {s.get('rpe','?')} | {dist_str}\n"
                f"    Structure: {s.get('structure','?')}\n"
                f"{contrainte_note}\n\n"
            )

    # ── Courses et J-X ──
    races_fmt = "\n".join(
        f"  {r['label']} ({r['date']}) : J-{r['jx']} | Objectif actuel: {r['sub_actuel']}"
        for r in races_jx
    )

    # ── Contexte zone calibration ──
    calib_fmt = ""
    if zone_calib.get("auto_update"):
        calib_fmt = (f"RECALIBRAGE ZONES DISPONIBLE : "
                     f"Z2 running → allure moy mesurée {zone_calib.get('proposed_z2_pace_str')} "
                     f"({zone_calib.get('reason')})")
    elif zone_calib.get("pending"):
        calib_fmt = f"Recalibrage zones : insuffisant ({zone_calib.get('pending_reason')})"

    # ═══ CONSTRUCTION PROMPT FINAL ═══════════════════════════════════════════

    system_prompt = f"""Tu es un entraîneur triathlon de niveau élite, spécialiste en périodisation pour Ironman, marathon et semi-marathon. Tu travailles avec Mattéo depuis plusieurs semaines et tu connais parfaitement son profil, ses données physiologiques et ses objectifs.

═══ PROFIL ATHLÈTE ═══
Prénom : Mattéo | Âge : 24 ans | Poids : 69 kg | Taille : 178 cm | IMC : 21.8

═══ ZONES FC ═══
{zones_fmt}

═══ OBJECTIFS ET ÉCHÉANCES ═══
{races_fmt}

═══ LACUNES DE DONNÉES — LIS AVANT TOUT ═══
Ces lacunes sont structurelles (matériel/API), pas des oublis de Mattéo :
- Vélo extérieur : AUCUNE donnée FC ni puissance (le Polar ne remonte pas les données du capteur vélo)
  → Ne jamais inférer une séance vélo comme "légère" faute de données — simplement la marquer sans FC
- Natation : pas de splits par longueur, uniquement distance totale + durée
- S1 à S7 : données physiologiques absentes (séances injectées manuellement)
- HRV : non disponible sur ce matériel
Si une donnée est absente pour une séance, INDIQUE explicitement "données indisponibles" plutôt que d'extrapoler

═══ RÈGLES D'ANALYSE — LIS ATTENTIVEMENT ═══

1. PRÉCISION ABSOLUE
   - Chaque affirmation doit être justifiée par un chiffre réel dans les données
   - Zéro généralité, zéro approximation — si la donnée n'est pas là, dis-le
   - Cite les dates, les km, les FC, les charges — jamais de vague

2. EFFICACITÉ AÉROBIE
   - Formule : pace_secondes ÷ fc_avg → PLUS LA VALEUR EST BASSE, MEILLEURE EST LA FORME
   - Exemple : 2.10 = excellent | 2.35 = dégradé | 3.00 = récupération/fatigue
   - Compare TOUJOURS avec le meilleur niveau mesuré fourni dans les métriques
   - Une valeur qui MONTE = régression. Une valeur qui BAISSE = progression

3. ATL / CTL / TSB
   - ATL = charge aiguë 7j (fatigue immédiate)
   - CTL = charge chronique 42j (forme de fond)
   - TSB = CTL - ATL (fraîcheur) : positif = frais, négatif = fatigué
   - Zone optimale de performance : TSB entre -10 et +5
   - Zone de surmenage : TSB < -30 → signal d'alarme

4. SUB CIBLES
   - Les valeurs dans "sub_proposes" sont PRÉ-CALCULÉES depuis l'allure Z2 réelle mesurée
   - Utilise-les comme base de départ
   - Ajuste uniquement si les données de la semaine le justifient clairement (ex: blessure, forme exceptionnelle)
   - Justifie tout changement avec un chiffre précis

5. PROGRAMME S+1 ET S+2 — STRUCTURE SEMAINE FIXE
   JOURS DE COURSE OBLIGATOIRES : Mardi, Jeudi, Vendredi, Dimanche (sortie longue)
   VÉLO + NATATION : Samedi uniquement
   LUNDI : TOUJOURS repos — ne jamais mettre de course le lundi, même en récupération active
   MERCREDI : repos ou natation uniquement — jamais de course
   NE JAMAIS afficher les jours de repos dans le programme (ni "Repos complet" ni "Récupération active")
   → Le programme ne doit contenir QUE les séances sportives réelles

   RÈGLE POST-COURSE (compétition ou course test) :
   Si la semaine analysée contient une course ou compétition :
   - Lundi S+1 → REPOS (ne pas afficher)
   - Mardi S+1 (J+3) → Z1-Z2 uniquement, volume normal autorisé, PAS de qualité
   - Jeudi S+1 (J+5) → Z2-Z3 max, PAS de tempo ni séance seuil ni fractionné Z4
   - Vendredi S+1 (J+6) → Z3 léger possible, PAS de Z4
   - Samedi/Dimanche → programme normal maintenu (vélo Z2, sortie longue non négociable)

   VOLUME : ne jamais descendre sous -10% du volume course habituel, même post-course
   Sortie longue du dimanche = NON NÉGOCIABLE, maintenir la distance prévue

   DÉPLACEMENT DE SÉANCE :
   Si tu déplaces une séance vers un jour déjà occupé → fusionner intelligemment, JAMAIS supprimer la séance existante
   Chaque jour ne peut avoir qu'une séance course maximum (vélo + natation peuvent s'additionner le samedi)

   RÈGLES GÉNÉRALES PROGRAMME :
   - Reprends EXACTEMENT les séances du programme fourni (dates, jours, sport, type, structure)
   - Conserve SANS EXCEPTION les valeurs de distance : dist_cible pour Course/Vélo, dist_cible_m pour Natation
   - Tu peux ajuster : zone, rpe, structure, duree — uniquement si la forme ou les contraintes le justifient
   - Si tu modifies quelque chose → remplis le champ "modification" avec la raison précise
   - Si contrainte calendrier (VOYAGE ou DEPART) sur un jour → déplace ou supprime la séance concernée

6. FORMAT STRICT
   - Sport natation : dist_cible = null, dist_cible_m = valeur entière en mètres (ex: 1600)
   - Sport course/vélo : dist_cible = valeur en km (ex: 12.0), dist_cible_m = null
   - Structure course : doit toujours se terminer par "— D+ : XXX m"
   - Tous les champs doivent être présents même si null
   - NE PAS inclure les séances de repos dans programme_s1 ou programme_s2

Tu dois produire UNIQUEMENT un JSON valide sans markdown, sans texte avant ou après.
Structure exacte attendue :

{{
  "diagnostic_semaine": "4-6 phrases précises avec chiffres réels. Analyse physiologique, pas de généralités.",
  "points_vigilance": ["point précis avec chiffres", "..."],
  "forme_globale": "EN_FORME | NORMAL | FATIGUE | SURCHARGE",
  "commentaire_forme": "1-2 phrases justifiées par TSB et efficacité aérobie",

  "programme_s1": [
    {{
      "date": "YYYY-MM-DD",
      "jour": "Lundi",
      "sport": "Course",
      "type": "type séance identique au programme",
      "duree": "1h05",
      "zone": "Z3 (146-160)",
      "rpe": "7",
      "structure": "15' EF → 3×10' tempo (r=2') → 10' RC — D+ : 130 m",
      "dist_cible": 11.0,
      "dist_cible_m": null,
      "modification": null
    }},
    {{
      "date": "YYYY-MM-DD",
      "jour": "Samedi",
      "sport": "Natation",
      "type": "Technique + éducatifs",
      "duree": "30min",
      "zone": "-",
      "rpe": "2",
      "structure": "4×50 facile → 11×50 éducatifs → 3×200 facile",
      "dist_cible": null,
      "dist_cible_m": 1600,
      "modification": null
    }}
  ],

  "programme_s2": [
    {{
      "date": "YYYY-MM-DD",
      "jour": "Lundi",
      "sport": "Course",
      "type": "type séance",
      "duree": "1h",
      "zone": "Z2",
      "rpe": "4",
      "structure": "EF continu — D+ : 80 m",
      "dist_cible": 10.0,
      "dist_cible_m": null,
      "modification": null
    }}
  ],

  "zones_proposees": null,
  "zones_confiance_haute": false,
  "zones_commentaire": "",

  "sub_proposes": {{
    "semi":     {{"sub_cible": "1:49:00", "allure_cible": "5:10/km", "probabilite_pct": 55, "note": "basé sur allure Z2 mesurée 5:12/km"}},
    "marathon": {{"sub_cible": "3:20:00", "allure_cible": "4:45/km", "probabilite_pct": 38, "note": ""}},
    "im703":    {{"sub_cible": "5:45:00", "allure_cible": null, "probabilite_pct": 55, "note": "données vélo/natation insuffisantes"}},
    "imfull":   {{"sub_cible": "11:30:00", "allure_cible": null, "probabilite_pct": 42, "note": ""}}
  }},

  "pending_updates": null
}}"""

    # ── Efficacité aérobie — référence historique ──
    all_efficiencies_hist = []
    for _, wd in [("", {"exercises": exercises_an})]:
        for ex in wd.get("exercises", []):
            if ex.get("aerobic_efficiency"):
                all_efficiencies_hist.append(ex["aerobic_efficiency"])
    # Chercher le meilleur dans les tendances
    best_eff_ref = sub_cibles.get("best_efficiency")
    current_eff  = sub_cibles.get("current_efficiency")
    eff_ratio    = sub_cibles.get("efficiency_ratio", 1.0)
    eff_context  = ""
    if best_eff_ref:
        eff_context = (
            f"RÉFÉRENCE EFFICACITÉ AÉROBIE (lower=better) :\n"
            f"  Meilleur niveau mesuré : {best_eff_ref} | Niveau actuel : {current_eff} | "
            f"Ratio actuel/meilleur : {eff_ratio} "
            f"({'✅ proche du meilleur' if eff_ratio < 1.05 else '⚠️ régression' if eff_ratio > 1.15 else '→ légère régression temporaire'})\n"
            f"  Z2 pace moy mesurée : {sub_cibles.get('z2_pace_avg','?')}/km "
            f"(sur {sub_cibles.get('nb_z2_sessions','?')} séances Z2 pures)"
        )

    # ── Sub cibles pré-calculées ──
    sub_fmt = "SUB CIBLES PRÉ-CALCULÉES (depuis données physiologiques réelles) :\n"
    if sub_cibles.get("erreur"):
        sub_fmt += f"  ⚠️ {sub_cibles['erreur']} — utiliser jugement clinique\n"
    else:
        for rid in ("semi", "marathon", "im703", "imfull"):
            sc = sub_cibles.get(rid, {})
            sub_fmt += (
                f"  {rid.upper()} : Sub {sc.get('sub_cible','?')} "
                f"@ {sc.get('allure_cible','—')}/km | "
                f"prob. {sc.get('probabilite_pct','?')}% | "
                f"base: {sc.get('base','?')}\n"
            )
        sub_fmt += "  → Ajuste ces valeurs si les données de la semaine le justifient (bonne/mauvaise forme, blessure...)"

    user_prompt = f"""ANALYSE SEMAINE {sem_label} — Aujourd'hui {today}

═══ MÉTRIQUES FITNESS ═══
ATL (fatigue aiguë 7j) : {atl_ctl['atl']}
CTL (forme chronique 42j) : {atl_ctl['ctl']}
TSB (fraîcheur) : {atl_ctl['tsb']} {'⚠️ FATIGUE' if atl_ctl['tsb'] < -20 else '✅ FRAIS' if atl_ctl['tsb'] > 5 else 'NEUTRE'}

{eff_context}

{sub_fmt}

═══ TENDANCES 4 SEMAINES PRÉCÉDENTES ═══
{trends_fmt}

═══ SÉANCES SEMAINE {sem_label} ═══
{exs_fmt}

═══ SOMMEIL SEMAINE ═══
{sleep_fmt}

═══ CORRÉLATION SOMMEIL J-1 / PERFORMANCE ═══
{corr_fmt or '  Données insuffisantes'}

═══ DELTA PRÉVU / RÉALISÉ ═══
Programme prévu cette semaine ({sem_label}) :
{delta_fmt or '  Programme non trouvé'}

═══ CALENDRIER EXACT S+1 — UTILISE CES DATES, NE LES CALCULE PAS TOI-MÊME ═══
⚠️ OBLIGATOIRE : copie ces dates exactes dans le JSON. Ne jamais deviner ou recalculer.
{cal_s1}

═══ PROGRAMME S+1 ({sem_next}) À REPROGRAMMER ═══
{prog_next_fmt or '  Programme non trouvé'}

═══ CALENDRIER EXACT S+2 — UTILISE CES DATES, NE LES CALCULE PAS TOI-MÊME ═══
⚠️ OBLIGATOIRE : copie ces dates exactes dans le JSON. Ne jamais deviner ou recalculer.
{cal_s2}

═══ PROGRAMME S+2 ({prog_next2.get("semaine", "S?") if prog_next2 else "S?"}) À REPROGRAMMER ═══
{prog_next2_fmt or '  Programme non trouvé'}

═══ CONTRAINTES CALENDRIER S+1 ET S+2 ═══
{json.dumps(constraints, ensure_ascii=False, indent=2)}

═══ {calib_fmt or 'Recalibrage zones : pas de données suffisantes'} ═══

═══ ALERTE POST-COURSE ═══
{"⚠️ COURSE/COMPÉTITION DÉTECTÉE cette semaine (" + race_date_str + ") — Appliquer OBLIGATOIREMENT la règle post-course pour S+1 : Mardi Z1-Z2 uniquement, Jeudi Z2-Z3 max (PAS de tempo/seuil/Z4), Vendredi Z3 léger max (PAS de Z4). Volume course maintenu ≥ -10% du volume habituel. Sortie longue dimanche NON NÉGOCIABLE." if has_race_this_week else "Aucune course détectée cette semaine — programme normal applicable."}

═══ CONTEXTE MANUEL ═══
{context_manuel or 'Aucun contexte particulier'}
"""

    return system_prompt, user_prompt


# ─── APPEL API CLAUDE ────────────────────────────────────────────────────────

def call_claude(system_prompt, user_prompt, api_key):
    import urllib.request
    payload = json.dumps({
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 4000,
        "system": system_prompt,
        "messages": [{"role": "user", "content": user_prompt}]
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "Content-Type":      "application/json",
            "x-api-key":         api_key,
            "anthropic-version": "2023-06-01",
        }
    )
    with urllib.request.urlopen(req, timeout=90) as resp:
        result = json.loads(resp.read())

    usage = result.get("usage", {})
    raw   = result["content"][0]["text"].strip()
    # Nettoyer éventuels backticks markdown
    raw   = re.sub(r'^```json\s*', '', raw)
    raw   = re.sub(r'```\s*$',     '', raw).strip()

    return json.loads(raw), usage


# ─── ÉCRITURE FICHIERS ────────────────────────────────────────────────────────

def write_json_atomic(path, data):
    """Écriture atomique avec fichier temporaire."""
    path = Path(path)
    tmp  = str(path) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, str(path))
    print(f"  ✅ Écrit : {path}")

def apply_coach_output(coach_json, prog_next_path, prog_next2_path, zones, races, zone_calib, atl_ctl=None, exercises_an=None, sleep_data=None, sub_cibles=None):
    """
    Applique le JSON du coach sur les fichiers :
    - semaine_S+1.json  : programme S+1 mis à jour directement
    - semaine_S+2.json  : programme S+2 mis à jour directement
    - races.json        : sub cibles + probabilités
    - hr_zones.json     : si confiance élevée
    - analysis_current.json : rapport complet
    """
    results = {"files_written": [], "pending_updates": None}

    # 1. Programme S+1
    if prog_next_path and prog_next_path.exists():
        prog_data = load_json(prog_next_path, {})
        new_seances = coach_json.get("programme_s1", [])
        if new_seances:
            # Merger : appliquer modifs coach sans ecraser seances manquantes
            orig = {s.get("jour","")+"_"+s.get("sport",""):s for s in prog_data.get("seances",[])}
            new_jours = {s.get("jour","")+"_"+s.get("sport","") for s in new_seances}
            merged = []
            for ns in new_seances:
                key = ns.get("jour","")+"_"+ns.get("sport","")
                base = dict(orig.get(key, {}))
                for k in ["description","coach_note","zones","rpe","duree_min","km"]:
                    if k in ns: base[k] = ns[k]
                for k in ["label","date","sport","jour","type"]:
                    if k in orig.get(key,{}) and k not in base: base[k] = orig[key][k]
                    elif k in ns and k not in base: base[k] = ns[k]
                merged.append(base)
            for os in prog_data.get("seances",[]):
                if os.get("jour","")+"_"+os.get("sport","") not in new_jours:
                    merged.append(os)
            prog_data["seances"] = merged
            prog_data["coach_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            write_json_atomic(prog_next_path, prog_data)
            results["files_written"].append(str(prog_next_path))

    # 2. Programme S+2
    if prog_next2_path and prog_next2_path.exists():
        prog_data2 = load_json(prog_next2_path, {})
        new_seances2 = coach_json.get("programme_s2", [])
        if new_seances2:
            orig2 = {s.get("jour","")+"_"+s.get("sport",""):s for s in prog_data2.get("seances",[])}
            new_jours2 = {s.get("jour","")+"_"+s.get("sport","") for s in new_seances2}
            merged2 = []
            for ns in new_seances2:
                key = ns.get("jour","")+"_"+ns.get("sport","")
                base = dict(orig2.get(key, {}))
                for k in ["description","coach_note","zones","rpe","duree_min","km"]:
                    if k in ns: base[k] = ns[k]
                for k in ["label","date","sport","jour","type"]:
                    if k in orig2.get(key,{}) and k not in base: base[k] = orig2[key][k]
                    elif k in ns and k not in base: base[k] = ns[k]
                merged2.append(base)
            for os in prog_data2.get("seances",[]):
                if os.get("jour","")+"_"+os.get("sport","") not in new_jours2:
                    merged2.append(os)
            prog_data2["seances"] = merged2
            prog_data2["coach_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            write_json_atomic(prog_next2_path, prog_data2)
            results["files_written"].append(str(prog_next2_path))

    # 2. races.json
    # Fusionner sub_proposes IA avec estimations calculées
    sub_proposes = coach_json.get("sub_proposes", {})
    if sub_cibles:
        for k in ["run_10k","semi","marathon",
                  "swim_900m","swim_1500m","swim_1900m","swim_3800m",
                  "bike_40km","bike_90km","bike_180km",
                  "im703","imfull"]:
            if k in sub_cibles:
                sub_proposes[k] = sub_cibles[k]
        if sub_cibles.get("data_quality"):
            sub_proposes["_data_quality"] = sub_cibles["data_quality"]
    if sub_proposes:
        races_updated = []
        for r in races:
            rid = r["id"]
            if rid in sub_proposes:
                sp = sub_proposes[rid]
                r["objectif"]       = f"Sub {sp['sub_cible']}"
                r["sub_cible"]      = sp["sub_cible"]
                r["allure_cible"]   = sp.get("allure_cible")
                r["probabilite_pct"] = sp.get("probabilite_pct")
                r["coach_note"]     = sp.get("note", "")
                r["updated_at"]     = datetime.now().strftime("%Y-%m-%d")
            races_updated.append(r)
        write_json_atomic(RACES_FILE, races_updated)
        results["files_written"].append(str(RACES_FILE))

    # 3. hr_zones.json (Option C)
    zones_proposees = coach_json.get("zones_proposees")
    confiance_haute = coach_json.get("zones_confiance_haute", False)
    if zones_proposees and confiance_haute:
        write_json_atomic(ZONES_FILE, zones_proposees)
        results["files_written"].append(str(ZONES_FILE))
    elif zones_proposees:
        results["pending_updates"] = {
            "type":    "zones",
            "data":    zones_proposees,
            "comment": coach_json.get("zones_commentaire", ""),
        }

    # 4. analysis_current.json
    analysis = {
        "generated_at":      datetime.now().isoformat(),
        "sem_label":         coach_json.get("sem_label", ""),
        "diagnostic":        coach_json.get("diagnostic_semaine", ""),
        "points_vigilance":  coach_json.get("points_vigilance", []),
        "forme_globale":     coach_json.get("forme_globale", ""),
        "commentaire_forme": coach_json.get("commentaire_forme", ""),
        "programme_s1":      coach_json.get("programme_s1", []),
        "programme_s2":      coach_json.get("programme_s2", []),
        "sub_proposes":      sub_proposes,
        "zones_proposees":   zones_proposees,
        "zones_confiance_haute": confiance_haute,
        "zones_commentaire": coach_json.get("zones_commentaire", ""),
        "pending_updates":   results.get("pending_updates"),
        "metriques": {
            "atl": (atl_ctl or {}).get("atl"),
            "ctl": (atl_ctl or {}).get("ctl"),
            "tsb": (atl_ctl or {}).get("tsb"),
            "volume_km": sum(e.get("distance_km",0) or 0 for e in (exercises_an or []) if e.get("sport","").upper()=="RUNNING"),
            "efficacite_moy": (summarise_week_exercises(exercises_an or []) or {}).get("efficacite_moy"),
            "efficacite_best": min((e["aerobic_efficiency"] for e in (exercises_an or []) if e.get("aerobic_efficiency")), default=None),
            "sleep_score_moyen": (sleep_data or {}).get("score_moyen"),
            "hr_nocturne_moy": (sleep_data or {}).get("fc_nocturne_moy"),
            "recharge_moyen": (sleep_data or {}).get("recharge_moyen"),
        },
        "files_written":     results["files_written"],
    }
    write_json_atomic(ANALYSIS_FILE, analysis)
    results["files_written"].append(str(ANALYSIS_FILE))

    return results


# ─── POINT D'ENTRÉE PRINCIPAL ─────────────────────────────────────────────────

def run(week_key_target=None, context_manuel="", dry_run=False):
    print("\n🏃 Coach IA — Démarrage de l'analyse\n")

    # ── Chargement de tous les fichiers ──
    print("📂 Chargement des données...")
    zones  = load_zones()
    races  = load_races()
    all_weeks = load_all_weeks()

    if not all_weeks:
        print("❌ Aucun fichier week_*.json trouvé dans", WEEKS_DIR)
        return None

    # ── Identifier la semaine à analyser ──
    if week_key_target:
        # Chercher la semaine précisée
        week_data = next(
            (wd for _, wd in all_weeks if wd.get("week_key") == week_key_target),
            None
        )
        if not week_data:
            print(f"❌ Semaine {week_key_target} introuvable")
            return None
    else:
        # Semaine courante : priorité week_current, sinon la plus récente
        wc_path = WEEKS_DIR / "week_current.json"
        if wc_path.exists():
            week_data = load_json(wc_path, {})
        else:
            _, week_data = all_weeks[-1]

    week_key  = week_data.get("week_key", "")
    sem_label = week_data.get("sem_label", "?")
    print(f"📅 Semaine analysée : {sem_label} ({week_key})")

    # ── Programme semaine analysée + S+1 + S+2 (Excel prioritaire) ──
    excel_programme = load_programme_from_excel()
    prog_current = {}
    prog_next    = {}
    prog_next2   = {}

    m_sem = re.search(r'(\d+)', str(sem_label))
    sem_num = int(m_sem.group(1)) if m_sem else None

    if excel_programme and sem_num:
        if sem_num in excel_programme:
            prog_current = {"seances": excel_programme[sem_num]}
        if sem_num + 1 in excel_programme:
            prog_next = {"semaine": f"Semaine {sem_num+1}",
                        "seances": excel_programme[sem_num + 1]}
        if sem_num + 2 in excel_programme:
            prog_next2 = {"semaine": f"Semaine {sem_num+2}",
                         "seances": excel_programme[sem_num + 2]}

    # Fallback sur fichiers JSON si Excel absent
    # On définit toujours les chemins fichiers (nécessaires pour l'écriture)
    next_num  = sem_num + 1 if sem_num else None
    next2_num = sem_num + 2 if sem_num else None
    prog_next_file  = PROG_DIR / f"semaine_{next_num:02d}.json"  if next_num  else None
    prog_next2_file = PROG_DIR / f"semaine_{next2_num:02d}.json" if next2_num else None

    if not prog_current:
        prog_current_file = find_programme_file(sem_label)
        prog_current = load_json(prog_current_file, {}) if prog_current_file else {}
    if not prog_next and prog_next_file and prog_next_file.exists():
        prog_next = load_json(prog_next_file, {})
    if not prog_next2 and prog_next2_file and prog_next2_file.exists():
        prog_next2 = load_json(prog_next2_file, {})

    # ── Contraintes calendrier S+1 et S+2 ──
    all_next_dates = []
    for prog in [prog_next, prog_next2]:
        for s in prog.get("seances", []):
            if s.get("date"): all_next_dates.append(s["date"])

    # ── Calendrier exact S+1 et S+2 (pour que le coach ne calcule pas les dates) ──
    JOURS_FR = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
    def build_week_calendar(monday_str):
        """Retourne le calendrier exact d'une semaine depuis son lundi."""
        try:
            monday = datetime.strptime(monday_str, "%Y-%m-%d")
            lines = []
            for i, jour in enumerate(JOURS_FR):
                d = monday + timedelta(days=i)
                lines.append(f"  {jour} = {d.strftime('%Y-%m-%d')}")
            return "\n".join(lines)
        except:
            return "  Dates non calculables"

    # Calculer les lundis de S+1 et S+2
    try:
        current_monday = datetime.strptime(week_key, "%Y-%m-%d")
        next_monday    = current_monday + timedelta(weeks=1)
        next2_monday   = current_monday + timedelta(weeks=2)
        cal_s1 = build_week_calendar(next_monday.strftime("%Y-%m-%d"))
        cal_s2 = build_week_calendar(next2_monday.strftime("%Y-%m-%d"))
    except:
        cal_s1 = "  Dates non calculables"
        cal_s2 = "  Dates non calculables"

    if all_next_dates:
        all_next_dates_sorted = sorted(all_next_dates)
        constraints = load_calendar_constraints(all_next_dates_sorted[0], all_next_dates_sorted[-1])
    else:
        try:
            next_monday = (datetime.strptime(week_key, "%Y-%m-%d") + timedelta(days=7)).strftime("%Y-%m-%d")
            next_sunday = (datetime.strptime(week_key, "%Y-%m-%d") + timedelta(days=20)).strftime("%Y-%m-%d")
            constraints = load_calendar_constraints(next_monday, next_sunday)
        except: constraints = {}

    # ── Pré-calculs ──
    print("🔢 Calculs en cours...")
    exercises_an = [analyse_exercise(ex, zones) for ex in week_data.get("exercises", [])]
    sleep_data   = analyse_sleep(week_data.get("sleep", []))
    sleep_corr   = correlate_sleep_perf(week_data.get("sleep", []), exercises_an)
    atl_ctl      = compute_atl_ctl_tsb(all_weeks)
    trends       = compute_trends(all_weeks, zones, week_key)
    delta        = compute_delta(prog_current.get("seances", []), exercises_an)
    zone_calib   = check_zone_recalibration(all_weeks, zones)
    # Construire analysis_data pour les probabilités
    _vol_km = sum(e.get('distance_km',0) or 0 for e in exercises_an if e.get('sport','').upper()=='RUNNING')
    # sleep_data peut être un dict résumé ou une liste
    if isinstance(sleep_data, dict):
        _hr_noc = sleep_data.get('fc_nocturne_moy') or 0
    elif sleep_data:
        _hr_noc = sum(s.get('heart_rate_avg',0) or 0 for s in sleep_data if isinstance(s,dict)) / max(1,len(sleep_data))
    else:
        _hr_noc = 0
    analysis_data = {
        'atl': atl_ctl.get('atl', 0),
        'ctl': atl_ctl.get('ctl', 0),
        'tsb': atl_ctl.get('tsb', 0),
        'volume_km': round(_vol_km, 2),
        'hr_nocturne_moy': round(_hr_noc, 1),
    }
    sub_cibles   = compute_sub_cibles(all_weeks, zones, analysis_data=analysis_data)

    print(f"  ATL={atl_ctl['atl']} CTL={atl_ctl['ctl']} TSB={atl_ctl['tsb']}")
    print(f"  {len(exercises_an)} séances | {len(trends)} semaines de tendances | "
          f"Zone calib: {'auto' if zone_calib.get('auto_update') else 'pending' if zone_calib.get('pending') else 'non'}")
    if sub_cibles.get("semi"):
        print(f"  Sub semi calculé : {sub_cibles['semi']['sub_cible']} "
              f"({sub_cibles['semi']['probabilite_pct']}%) | "
              f"Z2 moy : {sub_cibles.get('z2_pace_avg','?')}/km")

    # ── Construction du prompt ──
    system_prompt, user_prompt = build_prompt(
        week_data, exercises_an, sleep_data, sleep_corr,
        atl_ctl, trends, delta, prog_current, prog_next, prog_next2,
        zones, races, constraints, context_manuel, zone_calib,
        sub_cibles, cal_s1, cal_s2,
    )

    if dry_run:
        print("\n" + "="*60)
        print("SYSTEM PROMPT:\n", system_prompt[:2000], "...")
        print("\nUSER PROMPT:\n", user_prompt[:3000], "...")
        print(f"\nTaille approx: {len(system_prompt)+len(user_prompt)} chars")
        return None

    # ── Appel Claude Sonnet ──
    if not ANTHROPIC_KEY:
        print("❌ ANTHROPIC_API_KEY non définie")
        return None

    print("🤖 Appel Claude Sonnet...")
    try:
        coach_json, usage = call_claude(system_prompt, user_prompt, ANTHROPIC_KEY)
        print(f"  Tokens : {usage.get('input_tokens')} in / {usage.get('output_tokens')} out")
    except Exception as e:
        print(f"❌ Erreur API : {e}")
        return None

    coach_json["sem_label"] = sem_label

    # ── Application des résultats ──
    print("\n💾 Écriture des fichiers...")
    results = apply_coach_output(coach_json, prog_next_file, prog_next2_file, zones, races, zone_calib, atl_ctl=atl_ctl, exercises_an=exercises_an, sleep_data=sleep_data, sub_cibles=sub_cibles)

    # ── Résumé ──
    print(f"\n✅ Analyse terminée — {len(results['files_written'])} fichiers mis à jour")
    print(f"\n📊 DIAGNOSTIC :\n{coach_json.get('diagnostic_semaine','')}")
    print(f"\n⚡ FORME : {coach_json.get('forme_globale')} — {coach_json.get('commentaire_forme','')}")
    if coach_json.get("points_vigilance"):
        print("\n⚠️  VIGILANCE :")
        for p in coach_json["points_vigilance"]:
            print(f"  • {p}")

    return coach_json


# ─── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Coach IA Polar — Analyse hebdomadaire")
    parser.add_argument("--week",    type=str, default=None,
                        help="Semaine cible (ex: 2026-03-16)")
    parser.add_argument("--context", type=str, default="",
                        help="Contexte manuel (ex: 'douleur genou gauche mardi')")
    parser.add_argument("--dry-run", action="store_true",
                        help="Affiche le prompt sans appeler l'API")
    args = parser.parse_args()

    run(
        week_key_target=args.week,
        context_manuel=args.context,
        dry_run=args.dry_run,
    )
