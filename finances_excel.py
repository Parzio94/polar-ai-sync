"""
finances_excel.py — Import/Export Excel pour le module finances.
Format identique au template Fichier_financier_template.xlsx
"""
import io
from datetime import datetime

def export_finances_excel(data: dict) -> io.BytesIO:
    """Génère un fichier Excel au format du template."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl non installé : pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Styles ───────────────────────────────────────────────────────────────
    def hdr_style(ws, cell, txt):
        ws[cell] = txt
        ws[cell].font = Font(bold=True, color="FFFFFF", size=9)
        ws[cell].fill = PatternFill("solid", start_color="1A1A2E")
        ws[cell].alignment = Alignment(horizontal="center")

    def section_style(ws, cell, txt):
        ws[cell] = txt
        ws[cell].font = Font(bold=True, size=9, color="1565C0")
        ws[cell].fill = PatternFill("solid", start_color="E3F2FD")

    def total_style(ws, cell, val):
        ws[cell] = val
        ws[cell].font = Font(bold=True, size=9, color="2D6A4F")
        ws[cell].fill = PatternFill("solid", start_color="E8F5E9")
        ws[cell].number_format = '#,##0'

    def money(ws, cell, val):
        ws[cell] = float(val) if val else 0
        ws[cell].number_format = '#,##0'
        ws[cell].font = Font(size=9)

    def fmt_mois(m):
        """'2025-01' → datetime"""
        try: return datetime.strptime(m, "%Y-%m")
        except: return None

    # ── Feuille VIE ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Dépenses VIE")
    vie = data.get("vie", {})
    mois_vie = list((vie.get("revenus") or [{}])[0].get("mois", {}).keys())

    # En-tête
    ws["A2"] = "Les Revenus"; ws["B2"] = "Répartition"
    for i, m in enumerate(mois_vie):
        d = fmt_mois(m)
        if d: ws.cell(2, 3+i, d)
    ws.cell(2, 3+len(mois_vie), "Moyenne")
    ws.cell(2, 4+len(mois_vie), "Total")

    # Revenus
    row = 3
    for rev in (vie.get("revenus") or []):
        ws.cell(row, 1, rev.get("label", ""))
        for i, m in enumerate(mois_vie):
            money(ws, ws.cell(row, 3+i).coordinate, rev.get("mois", {}).get(m, 0))
        vals = [rev.get("mois", {}).get(m, 0) for m in mois_vie]
        ws.cell(row, 3+len(mois_vie)).value = sum(vals)/len(vals) if vals else 0
        ws.cell(row, 4+len(mois_vie)).value = sum(vals)
        row += 1

    # Total revenus
    for c in range(1, 5+len(mois_vie)):
        ws.cell(row, c, "")
    ws.cell(row, 1, "Total revenue")
    for i, m in enumerate(mois_vie):
        total = sum(r.get("mois", {}).get(m, 0) for r in (vie.get("revenus") or []))
        total_style(ws, ws.cell(row, 3+i).coordinate, total)
    row += 2

    # Loyer
    section_style(ws, f"A{row}", "Le LOYER"); row += 1
    ws.cell(row, 1, "Loyer")
    loyer_mois = (vie.get("loyer") or {}).get("mois", {})
    for i, m in enumerate(mois_vie):
        money(ws, ws.cell(row, 3+i).coordinate, loyer_mois.get(m, 0))
    row += 2

    # Charges
    section_style(ws, f"A{row}", "Les charges"); row += 1
    for chg in (vie.get("charges") or []):
        ws.cell(row, 1, chg.get("label", ""))
        for i, m in enumerate(mois_vie):
            money(ws, ws.cell(row, 3+i).coordinate, chg.get("mois", {}).get(m, 0))
        row += 1

    # Total charges
    ws.cell(row, 1, "Total des charges")
    for i, m in enumerate(mois_vie):
        total = sum(c.get("mois", {}).get(m, 0) for c in (vie.get("charges") or []))
        total_style(ws, ws.cell(row, 3+i).coordinate, total)
    row += 2

    # Épargne
    section_style(ws, f"A{row}", "Débit le 30 du mois"); row += 1
    ws.cell(row, 1, "Théorie de l'épargne")
    ep_th = {}
    for m in mois_vie:
        rev = sum(r.get("mois", {}).get(m, 0) for r in (vie.get("revenus") or []))
        chg = sum(c.get("mois", {}).get(m, 0) for c in (vie.get("charges") or []))
        loyer = loyer_mois.get(m, 0)
        ep_th[m] = rev - chg - loyer
        money(ws, ws.cell(row, 3+list(mois_vie).index(m)).coordinate, ep_th[m])
    row += 1

    ws.cell(row, 1, "Montant du débit bancaire")
    debit = (vie.get("debit_reel") or {})
    for i, m in enumerate(mois_vie):
        money(ws, ws.cell(row, 3+i).coordinate, debit.get(m, 0))
    row += 1

    ws.cell(row, 1, "Écart : réalité / théorie")
    for i, m in enumerate(mois_vie):
        ecart = ep_th.get(m, 0) - debit.get(m, 0)
        money(ws, ws.cell(row, 3+i).coordinate, ecart)
    row += 2

    # Ajuster largeurs
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    for i in range(len(mois_vie)):
        ws.column_dimensions[get_column_letter(3+i)].width = 10

    # ── Feuilles années ───────────────────────────────────────────────────────
    for annee in ["2024", "2025", "2026"]:
        an = (data.get("annees") or {}).get(annee, {})
        ws_a = wb.create_sheet(f"Dépenses {annee}")
        mois_an = list(((an.get("revenus") or [{}])[0]).get("mois", {}).keys())

        # En-tête
        ws_a["A3"] = "Les Revenus"
        for i, m in enumerate(mois_an):
            d = fmt_mois(m)
            if d: ws_a.cell(3, 2+i, d)
        ws_a.cell(3, 2+len(mois_an), "Moyenne")
        ws_a.cell(3, 3+len(mois_an), "Total")

        # Revenus
        row_a = 4
        for rev in (an.get("revenus") or []):
            ws_a.cell(row_a, 1, rev.get("label", ""))
            for i, m in enumerate(mois_an):
                money(ws_a, ws_a.cell(row_a, 2+i).coordinate, rev.get("mois", {}).get(m, 0))
            row_a += 1

        # Total revenus
        ws_a.cell(row_a, 1, "Total revenue")
        for i, m in enumerate(mois_an):
            total = sum(r.get("mois", {}).get(m, 0) for r in (an.get("revenus") or []))
            total_style(ws_a, ws_a.cell(row_a, 2+i).coordinate, total)
        row_a += 2

        # Charges
        section_style(ws_a, f"A{row_a}", "Les charges"); row_a += 1
        for chg in (an.get("charges") or []):
            ws_a.cell(row_a, 1, chg.get("label", ""))
            for i, m in enumerate(mois_an):
                money(ws_a, ws_a.cell(row_a, 2+i).coordinate, chg.get("mois", {}).get(m, 0))
            row_a += 1

        # Total charges
        ws_a.cell(row_a, 1, "Total des charges")
        for i, m in enumerate(mois_an):
            total = sum(c.get("mois", {}).get(m, 0) for c in (an.get("charges") or []))
            total_style(ws_a, ws_a.cell(row_a, 2+i).coordinate, total)
        row_a += 2

        # Épargne
        section_style(ws_a, f"A{row_a}", "Épargne"); row_a += 1
        ws_a.cell(row_a, 1, "Théorie de l'épargne")
        for i, m in enumerate(mois_an):
            rev = sum(r.get("mois", {}).get(m, 0) for r in (an.get("revenus") or []))
            chg = sum(c.get("mois", {}).get(m, 0) for c in (an.get("charges") or []))
            money(ws_a, ws_a.cell(row_a, 2+i).coordinate, rev - chg)
        row_a += 1
        ws_a.cell(row_a, 1, "Débit bancaire réel")
        for i, m in enumerate(mois_an):
            money(ws_a, ws_a.cell(row_a, 2+i).coordinate, (an.get("debit_reel") or {}).get(m, 0))
        row_a += 2

        # Investissements
        section_style(ws_a, f"A{row_a}", "Investissements"); row_a += 1
        inv = an.get("invest", {})
        ws_a.cell(row_a, 1, "Salaire brut annuel")
        ws_a.cell(row_a, 2, inv.get("salaire_brut", 0))
        row_a += 1
        ws_a.cell(row_a, 1, "Apport perso PEE")
        ws_a.cell(row_a, 2, inv.get("pee_apport_perso", 0))
        row_a += 1
        ws_a.cell(row_a, 1, "Apport CNP PEE")
        ws_a.cell(row_a, 2, inv.get("pee_apport_cnp", 0))
        row_a += 2

        ws_a.cell(row_a, 1, "Fonds PEE"); ws_a.cell(row_a, 2, "Répartition"); ws_a.cell(row_a, 3, "Apport perso"); ws_a.cell(row_a, 4, "Apport CNP")
        row_a += 1
        for f in (inv.get("pee_fonds") or []):
            ws_a.cell(row_a, 1, f.get("label", ""))
            ws_a.cell(row_a, 2, f.get("repartition", 0))
            ws_a.cell(row_a, 3, inv.get("pee_apport_perso", 0) * f.get("repartition", 0))
            ws_a.cell(row_a, 4, inv.get("pee_apport_cnp", 0) * f.get("repartition", 0))
            row_a += 1
        row_a += 1

        ws_a.cell(row_a, 1, "Livrets"); ws_a.cell(row_a, 2, "Taux"); ws_a.cell(row_a, 3, "Solde début"); ws_a.cell(row_a, 4, "Solde fin")
        row_a += 1
        for l in (inv.get("livrets") or []):
            ws_a.cell(row_a, 1, l.get("label", ""))
            ws_a.cell(row_a, 2, l.get("taux", 0))
            ws_a.cell(row_a, 3, l.get("solde_debut", 0))
            ws_a.cell(row_a, 4, l.get("solde_debut", 0) * (1 + l.get("taux", 0)))
            row_a += 1

        ws_a.column_dimensions["A"].width = 30
        for i in range(12):
            ws_a.column_dimensions[get_column_letter(2+i)].width = 11

    # ── Feuille simulation immo ───────────────────────────────────────────────
    ws_i = wb.create_sheet("Simulation Immo")
    sim = data.get("simulation_immo", {})

    ws_i["A1"] = "Simulation Immobilière"; ws_i["A1"].font = Font(bold=True, size=12)
    ws_i["A3"] = "Salaire mensuel net"; ws_i["B3"] = sim.get("salaire_mensuel", 0)
    ws_i["A4"] = "Remboursement navigo"; ws_i["B4"] = sim.get("remboursement_navigo", 0)
    ws_i["A5"] = "Trésorerie disponible"; ws_i["B5"] = sim.get("tresorie_disponible", 0)
    ws_i["A6"] = "Objectif prix achat"; ws_i["B6"] = sim.get("objectif_achat", 0)

    ws_i["A8"] = "Charges mensuelles propriétaire"
    row_i = 9
    for c in (sim.get("charges_propriete") or []):
        ws_i.cell(row_i, 1, c.get("label", ""))
        ws_i.cell(row_i, 2, c.get("montant", 0))
        row_i += 1

    # Classique
    ws_i.cell(row_i+1, 1, "EMPRUNT CLASSIQUE"); ws_i.cell(row_i+1, 1).font = Font(bold=True, color="1565C0")
    cl = sim.get("classique", {})
    ws_i.cell(row_i+2, 1, "Apport"); ws_i.cell(row_i+2, 2, cl.get("apport", 0))
    ws_i.cell(row_i+3, 1, "Taux annuel"); ws_i.cell(row_i+3, 2, cl.get("taux_annuel", 0))
    ws_i.cell(row_i+4, 1, "Durée (ans)"); ws_i.cell(row_i+4, 2, cl.get("duree_ans", 0))

    # PTZ
    ws_i.cell(row_i+1, 4, "EMPRUNT + PTZ"); ws_i.cell(row_i+1, 4).font = Font(bold=True, color="1565C0")
    pt = sim.get("ptz", {})
    ws_i.cell(row_i+2, 4, "Apport"); ws_i.cell(row_i+2, 5, pt.get("apport", 0))
    ws_i.cell(row_i+3, 4, "Montant PTZ"); ws_i.cell(row_i+3, 5, pt.get("montant_ptz", 0))
    ws_i.cell(row_i+4, 4, "Taux annuel"); ws_i.cell(row_i+4, 5, pt.get("taux_annuel", 0))
    ws_i.cell(row_i+5, 4, "Durée (ans)"); ws_i.cell(row_i+5, 5, pt.get("duree_ans", 0))

    ws_i.column_dimensions["A"].width = 28
    ws_i.column_dimensions["B"].width = 15
    ws_i.column_dimensions["D"].width = 28
    ws_i.column_dimensions["E"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_finances_excel(stream) -> dict:
    """
    Importe un fichier Excel (template ou export) et retourne le dict finances.
    Stratégie : lire les valeurs numériques des cellules et les mapper
    à la structure JSON par position/nom de ligne.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl non installé")

    wb = openpyxl.load_workbook(stream, data_only=True)
    from finances_module import default_finances
    data = default_finances()

    def safe_float(v):
        try: return float(v) if v is not None else 0
        except: return 0

    # Lire feuille VIE
    if "Dépenses VIE" in wb.sheetnames:
        ws = wb["Dépenses VIE"]
        vie = data["vie"]
        # Lire les mois depuis la ligne 2 (colonnes C à T = 3 à 20)
        mois_cols = {}
        for col in range(3, 22):
            cell = ws.cell(2, col).value
            if hasattr(cell, 'strftime'):
                mois_cols[col] = cell.strftime("%Y-%m")

        # Revenus lignes 3-6
        rev_labels = [r["id"] for r in vie["revenus"]]
        for idx, rid in enumerate(rev_labels):
            row = 3 + idx
            rev_obj = next((r for r in vie["revenus"] if r["id"] == rid), None)
            if rev_obj:
                for col, m in mois_cols.items():
                    if m in rev_obj["mois"]:
                        rev_obj["mois"][m] = safe_float(ws.cell(row, col).value)

        # Loyer ligne 10
        for col, m in mois_cols.items():
            if m in vie["loyer"]["mois"]:
                vie["loyer"]["mois"][m] = safe_float(ws.cell(10, col).value)

        # Charges lignes 13-23
        chg_ids = [c["id"] for c in vie["charges"]]
        for idx, cid in enumerate(chg_ids):
            row = 13 + idx
            chg_obj = next((c for c in vie["charges"] if c["id"] == cid), None)
            if chg_obj:
                for col, m in mois_cols.items():
                    if m in chg_obj["mois"]:
                        chg_obj["mois"][m] = safe_float(ws.cell(row, col).value)

        # Débit réel ligne 29
        for col, m in mois_cols.items():
            if m in vie["debit_reel"]:
                vie["debit_reel"][m] = safe_float(ws.cell(29, col).value)

    # Lire feuilles années
    for annee in ["2024", "2025", "2026"]:
        sheet_name = f"Dépenses {annee}"
        if sheet_name not in wb.sheetnames:
            continue
        ws_a = wb[sheet_name]
        an = data["annees"][annee]

        # Mois en ligne 3 (colonnes B à N)
        mois_cols_a = {}
        for col in range(2, 16):
            cell = ws_a.cell(3, col).value
            if hasattr(cell, 'strftime'):
                mois_cols_a[col] = cell.strftime("%Y-%m")

        # Revenus lignes 4-7
        for idx, rev in enumerate(an["revenus"]):
            row = 4 + idx
            for col, m in mois_cols_a.items():
                if m in rev["mois"]:
                    rev["mois"][m] = safe_float(ws_a.cell(row, col).value)

        # Charges lignes 12-16
        for idx, chg in enumerate(an["charges"]):
            row = 12 + idx
            for col, m in mois_cols_a.items():
                if m in chg["mois"]:
                    chg["mois"][m] = safe_float(ws_a.cell(row, col).value)

        # Débit réel ligne 22
        for col, m in mois_cols_a.items():
            if m in an["debit_reel"]:
                an["debit_reel"][m] = safe_float(ws_a.cell(22, col).value)

    return data
